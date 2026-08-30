"""
(주)대교통신 사내 HR 플랫폼 - Phase 1 (근태입력 셀프서비스 MVP)

실행:
    streamlit run app.py

기본 관리자 계정: admin / changeme123  (반드시 최초 로그인 후 비밀번호를 바꿀 직원용 계정을 새로 만들고,
                                       admin 계정 비밀번호도 바꿔서 사용하세요)
"""

import io
import re
from calendar import Calendar
from datetime import date, timedelta

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from streamlit_sortables import sort_items

import db
import ai_review
import excel_export

st.set_page_config(page_title="(주)대교통신 HR 플랫폼", page_icon="🗂️", layout="wide")
db.init_db()

def _build_bulk_template():
    """직원 일괄 등록용 빈 엑셀 양식 (예시 3줄 포함) 생성"""
    sample = pd.DataFrame([
        {"아이디": "D100001", "이름": "홍길동", "구분": "본사", "부서/매장": "관리팀",
         "초기비밀번호": "changeme01", "권한": "employee"},
        {"아이디": "D100002", "이름": "김철수", "구분": "직영", "부서/매장": "범계역직영점",
         "초기비밀번호": "changeme02", "권한": "employee"},
        {"아이디": "D100003", "이름": "이영희", "구분": "소사장", "부서/매장": "대교대리점 군포역점",
         "초기비밀번호": "changeme03", "권한": "employee"},
    ])
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        sample.to_excel(writer, index=False, sheet_name="직원목록")
    return buf.getvalue()


# ---------------------------------------------------------------------------
# 탭 순서 (관리자가 ▲▼ 버튼으로 자유롭게 바꿀 수 있도록 DB 설정에 저장)
# ---------------------------------------------------------------------------
TAB_LABELS = {
    "ai": "카톡 AI 검토",
    "export": "엑셀 내보내기",
    "bulk": "근태일괄입력",
    "emp": "근태개별입력",
    "status": "전체 근태현황",
    "kakao": "카톡 근태 가져오기",
    "calendar": "반영 현황 캘린더",
    "admin": "직원 계정관리",
}
DEFAULT_TAB_ORDER = ["ai", "calendar", "bulk", "emp", "status", "export", "kakao", "admin"]
_TAB_ORDER_SETTING_KEY = "admin_tab_order"


def _get_tab_order():
    raw = db.get_setting(_TAB_ORDER_SETTING_KEY, "")
    order = [k for k in raw.split(",") if k in TAB_LABELS] if raw else []
    for k in DEFAULT_TAB_ORDER:
        if k not in order:
            order.append(k)
    return order


def _save_tab_order(order):
    db.set_setting(_TAB_ORDER_SETTING_KEY, ",".join(order))


def tab_order_settings_view():
    st.caption(
        "아래 목록을 손가락으로 눌러서 위아래로 끌어놓으면(드래그) 순서가 바로 저장되고, "
        "탭 순서도 그대로 바뀌어요. (Streamlit 특성상 화면 위쪽 탭 글자 자체를 직접 끌 수는 없어서, "
        "이 목록을 대신 드래그하는 방식이에요.)"
    )
    order = _get_tab_order()
    current_labels = [TAB_LABELS[k] for k in order]
    new_labels = sort_items(current_labels, direction="vertical", key="tab_order_sortable")
    if new_labels != current_labels:
        label_to_key = {v: k for k, v in TAB_LABELS.items()}
        new_order = [label_to_key[l] for l in new_labels]
        _save_tab_order(new_order)
        st.rerun()


# ---------------------------------------------------------------------------
# 로그인 / 세션 관리
# ---------------------------------------------------------------------------
def login_view():
    st.title("🗂️ (주)대교통신 HR 플랫폼")
    st.caption("근태입력 셀프서비스 (Phase 1)")

    with st.form("login_form"):
        username = st.text_input("아이디")
        password = st.text_input("비밀번호", type="password")
        submitted = st.form_submit_button("로그인", use_container_width=True)

    if submitted:
        user = db.verify_user(username.strip(), password)
        if user:
            st.session_state["user"] = user
            st.rerun()
        else:
            st.error("아이디 또는 비밀번호가 올바르지 않습니다.")


def logout_button():
    with st.sidebar:
        user = st.session_state["user"]
        st.markdown(f"**{user['name']}**")
        st.caption(f"{user['category']} · {user['department']}")
        st.caption("관리자" if user["role"] == "admin" else "직원")
        if st.button("로그아웃", use_container_width=True):
            del st.session_state["user"]
            st.rerun()


# ---------------------------------------------------------------------------
# 직원용: 근태입력
# ---------------------------------------------------------------------------
def employee_view(user):
    st.header("근태 입력")

    col1, col2 = st.columns([1, 1])
    with col1:
        with st.form("attendance_form"):
            work_date = st.date_input("날짜", value=date.today())
            code = st.selectbox("근태 코드", db.ATTENDANCE_CODES)
            memo = st.text_input("메모 (선택)", placeholder="예: 오전 반차, 결혼식 참석 등")
            submitted = st.form_submit_button("입력/수정 저장", use_container_width=True)
        if submitted:
            db.upsert_attendance(user["id"], work_date.isoformat(), code, memo)
            st.success(f"{work_date.isoformat()} 근태가 '{code}'(으)로 저장되었습니다.")
            st.rerun()

    with col2:
        st.markdown("**근태 코드 안내**")
        st.caption(
            "정상출근 · 연차 · 반차 · 지각 · 조퇴 · 특근 · 당직 · 교육 · "
            "경조 · 예비군 · 무급 · 개인용무 · 휴무 · 기타\n\n"
            "같은 날짜에 다시 입력하면 기존 내용을 덮어씁니다(수정)."
        )
        st.info(
            "💡 담당자가 '일괄입력'에서 같은 날짜를 저장하면 여기서 입력한 내용도 "
            "그 내용으로 덮어써요. 평소엔 담당자의 일괄입력을 기본으로 하고, "
            "개인 입력은 급하게 미리 남겨둘 때만 쓰는 걸 추천해요."
        )

    st.divider()
    st.subheader("내 최근 입력 내역")

    default_start = date.today() - timedelta(days=30)
    c1, c2 = st.columns(2)
    with c1:
        start = st.date_input("조회 시작일", value=default_start, key="emp_start")
    with c2:
        end = st.date_input("조회 종료일", value=date.today(), key="emp_end")

    records = db.get_user_attendance(user["id"], start.isoformat(), end.isoformat())
    if not records:
        st.info("해당 기간에 입력된 근태 내역이 없습니다.")
        return

    for r in records:
        with st.container(border=True):
            top_l, top_r = st.columns([2, 2])
            top_l.markdown(f"**{r['work_date']}**")
            top_r.caption(f"최종수정 {(r['updated_at'] or '')[:16].replace('T', ' ')}")

            code_idx = db.ATTENDANCE_CODES.index(r["code"]) if r["code"] in db.ATTENDANCE_CODES else 0
            c_code, c_memo = st.columns([1, 2])
            new_code = c_code.selectbox(
                "근태코드", db.ATTENDANCE_CODES, index=code_idx,
                key=f"hist_code_{r['work_date']}", label_visibility="collapsed",
            )
            new_memo = c_memo.text_input(
                "메모", value=r["memo"] or "", key=f"hist_memo_{r['work_date']}",
                label_visibility="collapsed", placeholder="메모",
            )

            b_edit, b_del = st.columns(2)
            if b_edit.button("수정 저장", key=f"hist_edit_{r['work_date']}", use_container_width=True):
                db.upsert_attendance(user["id"], r["work_date"], new_code, new_memo)
                st.success(f"{r['work_date']} 근태가 수정되었습니다.")
                st.rerun()
            if b_del.button("삭제", key=f"hist_del_{r['work_date']}", type="secondary", use_container_width=True):
                db.delete_attendance(user["id"], r["work_date"])
                st.success(f"{r['work_date']} 근태 입력이 삭제되었습니다.")
                st.rerun()


# ---------------------------------------------------------------------------
# 일괄입력: 본사 부서 / 직영·소사장 매장 담당자가 하루치 소속 인원 전체를 한번에 입력
# ---------------------------------------------------------------------------
def bulk_entry_view(user):
    st.header("일괄입력")
    st.caption("본사 부서, 직영·소사장 매장 담당자가 하루 단위로 소속 인원 전체 근태를 한번에 입력할 수 있어요.")
    st.caption(
        "💡 헷갈리지 않으려면: 이 화면을 '공식 입력'으로 쓰고, 아직 선택 안 한 사람은 "
        "그대로 두면 저장되지 않아요 — 정상출근이 자동으로 저장되지 않으니 안심하고 "
        "빈 채로 넘어가도 됩니다."
    )

    c1, c2, c3 = st.columns(3)
    with c1:
        work_date = st.date_input("날짜", value=date.today(), key="bulk_date")
    with c2:
        cat_options = db.CATEGORIES
        default_cat_idx = cat_options.index(user["category"]) if user["category"] in cat_options else 0
        category = st.selectbox("구분", cat_options, index=default_cat_idx, key="bulk_category")
    with c3:
        org_options = db.get_org_units(category)
        if category == user["category"] and user["department"] in org_options:
            default_org_idx = org_options.index(user["department"])
        else:
            default_org_idx = 0
        org_unit = st.selectbox(
            "부서/매장 선택", org_options, index=default_org_idx, key=f"bulk_org_{category}"
        )

    members = db.list_users_by_org(category, org_unit)
    if not members:
        st.info(f"'{org_unit}'에 소속된 직원이 없습니다.")
        return

    existing = db.get_attendance_for_date(work_date.isoformat(), category, org_unit)

    entered_count = len(existing)
    member_count = len(members)
    if work_date == date.today():
        if entered_count == 0:
            st.warning(f"⚠️ 오늘({work_date.isoformat()}) {org_unit} 근태를 아직 아무도 입력하지 않았어요.")
        elif entered_count < member_count:
            st.info(f"📝 오늘 {org_unit} {entered_count}/{member_count}명 입력 완료. 나머지 인원도 확인해주세요.")
        else:
            st.success(f"✅ 오늘 {org_unit} 전원({member_count}명) 입력 완료!")
    else:
        st.caption(f"선택한 날짜({work_date.isoformat()}) 기준 {org_unit} {entered_count}/{member_count}명 입력됨")

    with st.expander(f"{org_unit} 이번 달 날짜별 입력 현황 (언제 안 올렸는지 확인)"):
        month_start = work_date.replace(day=1)
        today = date.today()
        month_end = min(
            date(work_date.year, work_date.month + 1, 1) - timedelta(days=1)
            if work_date.month < 12 else date(work_date.year, 12, 31),
            today,
        )
        counts = db.get_daily_entry_counts(
            category, org_unit, month_start.isoformat(), month_end.isoformat()
        )
        weekday_kr = ["월", "화", "수", "목", "금", "토", "일"]
        rows = []
        d = month_start
        while d <= month_end:
            cnt = counts.get(d.isoformat(), 0)
            if cnt == 0:
                status = "❌ 미입력"
            elif cnt < member_count:
                status = f"⚠️ 일부만 ({cnt}/{member_count}명)"
            else:
                status = f"✅ 완료 ({cnt}/{member_count}명)"
            rows.append({"날짜": d.isoformat(), "요일": weekday_kr[d.weekday()], "상태": status})
            d += timedelta(days=1)
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    NOT_SELECTED = "─ 선택 안 함 ─"
    BULK_CODE_OPTIONS = [NOT_SELECTED] + db.ATTENDANCE_CODES

    with st.form("bulk_attendance_form"):
        entries = {}
        for m in members:
            has_prev = m["id"] in existing
            prev = existing.get(m["id"], {})
            prev_code = prev.get("code", "")
            prev_memo = prev.get("memo", "")
            code_idx = BULK_CODE_OPTIONS.index(prev_code) if has_prev and prev_code in db.ATTENDANCE_CODES else 0

            col_name, col_code, col_memo = st.columns([2, 2, 3])
            with col_name:
                st.markdown(f"**{m['name']}**")
                if has_prev:
                    st.caption(f"✅ 입력됨 ({prev_code})")
                else:
                    st.caption("⏳ 미입력")
            with col_code:
                code = st.selectbox(
                    "근태 코드", BULK_CODE_OPTIONS, index=code_idx,
                    key=f"bulk_code_{m['id']}", label_visibility="collapsed",
                )
            with col_memo:
                memo = st.text_input(
                    "메모", value=prev_memo, key=f"bulk_memo_{m['id']}",
                    label_visibility="collapsed", placeholder="메모 (선택)",
                )
            entries[m["id"]] = (code, memo, m["name"])

        submitted = st.form_submit_button(f"{org_unit} 전체 저장 ({work_date.isoformat()})", use_container_width=True)

    if submitted:
        saved = 0
        skipped_names = []
        for user_id, (code, memo, name) in entries.items():
            if code == NOT_SELECTED:
                skipped_names.append(name)
                continue
            db.upsert_attendance(user_id, work_date.isoformat(), code, memo)
            saved += 1
        if saved:
            st.success(f"{work_date.isoformat()} 기준 {org_unit} {saved}명 근태가 저장되었습니다.")
        if skipped_names:
            st.warning(
                f"⏳ 근태를 선택하지 않아 저장하지 않은 인원 {len(skipped_names)}명: "
                f"{', '.join(skipped_names)} — 확인 후 다시 입력해주세요."
            )
        st.rerun()


# ---------------------------------------------------------------------------
# 카톡 근태 가져오기: 본사/직영/소사장 카톡방에 매일 올라오는 근태 메시지를
# 그대로 붙여넣으면 근태표 형식으로 파싱 (직원들의 카톡 습관은 그대로 유지)
# ---------------------------------------------------------------------------

_KAKAO_STATUS_MAP = {
    "출근": "정상출근",
    "연차": "연차",
    "반차": "반차",
    "오전반차": "반차",
    "오후반차": "반차",
    "지각": "지각",
    "조퇴": "조퇴",
    "특근": "특근",
    "당직": "당직",
    "교육": "교육",
    "경조": "경조",
    "예비군": "예비군",
    "무급": "무급",
    "개인용무": "개인용무",
    "휴무": "휴무",
}

_DATE_DIVIDER_RE = re.compile(r"(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일")

# 카카오톡 "대화 내보내기" .txt 파일은 메시지 첫 줄이 "[보낸사람] [오전/오후 시:분] 내용" 형태로 시작함.
# 화면에서 복사해서 붙여넣을 땐 이 접두어가 없을 수도 있어서, 있으면 제거하고 없으면 그대로 둠.
_SENDER_PREFIX_RE = re.compile(r"^\[.+?\]\s*\[(?:오전|오후)\s*\d{1,2}:\d{2}\]\s*")

_STATUS_TIME_MEMO = {"오전반차": "오전", "오후반차": "오후"}

# 파서가 어떤 패턴으로도 못 잡았지만, 콜론이 있거나 근태 관련 단어가 들어있어서
# 혹시 놓친 근태 보고일 수도 있는 줄을 "확인이 필요할 수도 있는 줄"로 따로 모아서 보여줌.
_MAYBE_ATTENDANCE_RE = re.compile(r"[:：]|출근|휴무|연차|반차|지각|조퇴|특근|당직|퇴근")

# 카카오톡 "대화 내보내기" .txt 파일 맨 위에 항상 붙는 안내 줄(대화방 이름, 저장한 날짜)은 근태와 무관하니 무시
_EXPORT_METADATA_RE = re.compile(r"^저장한 날짜\s*[:：]|님과\s*카카오톡\s*대화\s*$")


def _strip_kakao_prefix(raw_line):
    return _SENDER_PREFIX_RE.sub("", raw_line, count=1).strip()


def parse_headquarters_chat(text, fallback_date):
    """본사 카톡 형식: "8월 28일 (금) 전산팀 출근현황" 헤더 다음 "이름 : 상태" 줄들.
    "-이름 : 상태"처럼 앞에 - 가 붙는 경우, 카카오톡 대화 내보내기 [보낸사람] [시간] 접두어도 처리.
    반환값: (인식된 근태 목록, 확인이 필요할 수도 있는 줄 목록)"""
    header_re = re.compile(r"(\d{1,2})월\s*(\d{1,2})일?.*?(?:출근현황|근태)")
    name_status_re = re.compile(r"^\s*-?\s*([가-힣A-Za-z0-9]+)\s*[:：]\s*(.+?)\s*$")

    results = []
    unrecognized = []
    current_date = fallback_date
    for raw in text.splitlines():
        line = _strip_kakao_prefix(raw)
        if not line or _EXPORT_METADATA_RE.search(line):
            continue
        dm = _DATE_DIVIDER_RE.search(line)
        if dm:
            try:
                current_date = date(int(dm.group(1)), int(dm.group(2)), int(dm.group(3)))
            except ValueError:
                pass
            continue
        hm = header_re.search(line)
        if hm:
            month, day = int(hm.group(1)), int(hm.group(2))
            try:
                current_date = date(current_date.year, month, day)
            except ValueError:
                pass
            continue
        nm = name_status_re.match(line)
        if nm:
            name, status_text = nm.group(1), nm.group(2).strip()
            code = _KAKAO_STATUS_MAP.get(status_text)
            results.append({
                "raw": line, "name": name, "store": None,
                "work_date": current_date, "code": code or "기타",
                "memo": _STATUS_TIME_MEMO.get(status_text, "" if code else status_text),
            })
            continue
        if _MAYBE_ATTENDANCE_RE.search(line):
            unrecognized.append((line, current_date))
    return results, unrecognized


def parse_jikyeong_chat(text, fallback_date):
    """직영 카톡 형식: "매장명 출근보고" / "10시 이름1 이름2" / "휴무 이름A 이름B" / "이상입니다".
    반환값: (인식된 근태 목록, 확인이 필요할 수도 있는 줄 목록)"""
    report_start_re = re.compile(r"^(.+?)\s*출근보고\s*$")
    time_line_re = re.compile(r"^\d{1,2}시\s*(.+)$")
    end_re = re.compile(r"^이상입니다\s*$")

    lines = [_strip_kakao_prefix(l) for l in text.splitlines()]
    results = []
    unrecognized = []
    current_date = fallback_date
    i = 0
    while i < len(lines):
        line = lines[i]
        if not line or _EXPORT_METADATA_RE.search(line):
            i += 1
            continue
        rm = report_start_re.match(line)
        if rm:
            store = rm.group(1).strip()
            block = {}
            i += 1
            if i < len(lines):
                tm = time_line_re.match(lines[i])
                if tm:
                    for nm in tm.group(1).split():
                        block[nm] = "정상출근"
                    i += 1
            while i < len(lines) and not end_re.match(lines[i]) and not report_start_re.match(lines[i]):
                parts = lines[i].split()
                if parts and parts[0] in _KAKAO_STATUS_MAP:
                    code = _KAKAO_STATUS_MAP[parts[0]]
                    for nm in parts[1:]:
                        block[nm] = code
                elif lines[i] and _MAYBE_ATTENDANCE_RE.search(lines[i]):
                    unrecognized.append((lines[i], current_date))
                i += 1
            if i < len(lines) and end_re.match(lines[i]):
                i += 1
            for name, code in block.items():
                results.append({
                    "raw": f"{store} 출근보고", "name": name, "store": store,
                    "work_date": current_date, "code": code, "memo": "",
                })
            continue
        dm = _DATE_DIVIDER_RE.search(line)
        if dm:
            try:
                current_date = date(int(dm.group(1)), int(dm.group(2)), int(dm.group(3)))
            except ValueError:
                pass
            i += 1
            continue
        if _MAYBE_ATTENDANCE_RE.search(line):
            unrecognized.append((line, current_date))
        i += 1
    return results, unrecognized


def parse_sosajang_chat(text, fallback_date):
    """소사장 카톡 형식: "매장명 출근입니다" / "매장명 휴무입니다 매장은 CLOSE 입니다" 등 한 줄 자기 보고.
    "금일 휴무입니다"처럼 메시지에 매장명이 없으면, 바로 위 발신자 줄("대교_매장명 이름소사장")에서
    매장명을 찾아 이어서 사용."""
    msg_re = re.compile(r"^(.+?)\s*(출근|휴무|퇴근)입니다")
    short_stores = [s.replace("대교대리점 ", "") for s in db.SOSAJANG_STORES]
    no_store_words = {"금일", "오늘", "익일", "내일"}

    results = []
    unrecognized = []
    current_date = fallback_date
    current_store = None
    for raw in text.splitlines():
        line = _strip_kakao_prefix(raw)
        if not line or _EXPORT_METADATA_RE.search(line):
            continue

        store_matches = [s for s in short_stores if s in line]
        found_store = max(store_matches, key=len) if store_matches else None

        mm = msg_re.match(line)
        if mm:
            body_store = mm.group(1).strip()
            if found_store:
                store = found_store
            elif body_store in no_store_words:
                store = current_store or body_store
            else:
                store = body_store
            status = mm.group(2)
            code = "휴무" if status == "휴무" else "정상출근"
            memo = "CLOSE" if "CLOSE" in line.upper() else ""
            results.append({
                "raw": line, "name": None, "store": store,
                "work_date": current_date, "code": code, "memo": memo,
            })
            if found_store:
                current_store = found_store
            continue

        dm = _DATE_DIVIDER_RE.search(line)
        if dm:
            try:
                current_date = date(int(dm.group(1)), int(dm.group(2)), int(dm.group(3)))
            except ValueError:
                pass
            continue

        if found_store:
            current_store = found_store
        elif _MAYBE_ATTENDANCE_RE.search(line):
            unrecognized.append((line, current_date))
    return results, unrecognized


_KAKAO_EXAMPLES = {
    "본사": "8월 28일 (금) 전산팀 출근현황\n최요안나 : 출근\n동희원 : 연차\n김선영 : 휴무",
    "직영": "상동점 출근보고\n10시 유경학\n연차 김찬양\n이상입니다",
    "소사장": "군포역점 출근입니다\n인덕원점 휴무입니다 매장은 CLOSE 입니다",
}


def _decode_uploaded_text(uploaded_file):
    raw_bytes = uploaded_file.getvalue()
    for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            return raw_bytes.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    return raw_bytes.decode("utf-8", errors="ignore")


def _kakao_calendar_view(category, year, month):
    """선택한 구분(본사/직영/소사장)·연/월 기준, 날짜별로 카톡 근태가 이미 반영됐는지 한눈에 보여주는 달력.
    ✅ 등록된 인원 전원 반영 / ⚠️ 일부만 반영 / ❌ 아무도 반영 안 됨 / 회색 = 아직 지나지 않은 날짜(판단 안 함)."""
    month_start = date(year, month, 1)
    month_end = date(year, 12, 31) if month == 12 else date(year, month + 1, 1) - timedelta(days=1)
    today = date.today()

    if category == "소사장":
        # 소사장은 '사람'이 아니라 '매장' 단위로 근태를 관리하므로 매장 기준으로 센다
        cat_employees = [{"id": s, "name": s} for s in db.SOSAJANG_STORES]
        total = len(cat_employees)
        by_date = {}
        for r in db.get_store_attendance(month_start.isoformat(), month_end.isoformat()):
            by_date.setdefault(r["work_date"], set()).add(r["store"])
        unit_label = f"매장 {total}곳"
    else:
        cat_employees = [e for e in db.list_users(include_inactive=False) if e["category"] == category]
        total = len(cat_employees)
        if total == 0:
            st.caption(f"등록된 {category} 직원이 없어 반영 현황을 표시할 수 없어요.")
            return
        records = db.get_attendance_records_for_matrix(
            category, month_start.isoformat(), month_end.isoformat()
        )
        by_date = {}
        for r in records:
            by_date.setdefault(r["work_date"], set()).add(r["user_id"])
        unit_label = f"인원 {total}명"

    st.caption(
        f"등록된 {category} {unit_label} 기준 — ✅ 전원 반영 / ⚠️ 일부만 반영 / ❌ 미반영 "
        f"/ 빈 칸은 아직 지나지 않은 날짜예요."
    )

    # st.columns는 모바일 좁은 화면에서 자동으로 세로로 쌓여버려서 달력이 깨짐 ->
    # HTML table 하나로 통째로 그려서 항상 7칸 가로 배치가 유지되게 함
    weekday_kr = ["월", "화", "수", "목", "금", "토", "일"]
    header_cells = "".join(
        f"<th style='padding:2px 0;font-size:11px;color:#888;font-weight:normal'>{wd}</th>" for wd in weekday_kr
    )
    rows_html = [f"<tr>{header_cells}</tr>"]

    for week in Calendar(firstweekday=0).monthdatescalendar(year, month):
        cells = []
        for day in week:
            if day.month != month:
                cells.append("<td></td>")
            elif day > today:
                cells.append(
                    "<td style='text-align:center;border:1px solid #eee;border-radius:6px;"
                    "padding:3px 0;color:#ccc'>"
                    f"<div style='font-size:11px'>{day.day}</div><div style='font-size:13px'>&nbsp;</div></td>"
                )
            else:
                cnt = len(by_date.get(day.isoformat(), set()))
                icon = "❌" if cnt == 0 else ("⚠️" if cnt < total else "✅")
                cells.append(
                    "<td style='text-align:center;border:1px solid #eee;border-radius:6px;padding:3px 0'>"
                    f"<div style='font-size:11px'>{day.day}</div><div style='font-size:13px'>{icon}</div></td>"
                )
        rows_html.append(f"<tr>{''.join(cells)}</tr>")

    table_html = (
        "<table style='width:100%;border-collapse:separate;border-spacing:2px;table-layout:fixed'>"
        + "".join(rows_html) + "</table>"
    )
    st.markdown(table_html, unsafe_allow_html=True)

    weekday_kr2 = ["월", "화", "수", "목", "금", "토", "일"]
    emp_names_by_id = {e["id"]: e["name"] for e in cat_employees}
    missing_lines = []
    d = month_start
    while d <= min(month_end, today):
        entered_ids = by_date.get(d.isoformat(), set())
        missing_ids = [uid for uid in emp_names_by_id if uid not in entered_ids]
        if missing_ids:
            missing_names = ", ".join(
                sorted(emp_names_by_id[uid] for uid in missing_ids)
            )
            missing_lines.append(f"**{d.month}/{d.day}({weekday_kr2[d.weekday()]})** — {missing_names}")
        d += timedelta(days=1)

    # 예전엔 펼침 메뉴 안에 숨겨뒀는데, "캘린더 안에 이름이 바로 보였으면 좋겠다"는 요청으로
    # 클릭 없이 항상 보이게 바꿈
    who = "매장" if category == "소사장" else "사람"
    st.markdown(f"**❌⚠️ 날짜별로 아직 안 올라온 {who} ({len(missing_lines)}일)**")
    if not missing_lines:
        st.caption(f"이번 달은 지난 날짜 전부 등록된 {who} 전체가 반영됐어요!")
    else:
        for line in missing_lines:
            st.markdown(line)


def kakao_calendar_view(user):
    st.subheader("반영 현황 캘린더")
    st.caption(
        "구분·연/월을 고르면 이번 달에 어느 날짜는 카톡 근태를 이미 가져왔고, 어느 날짜는 아직인지 "
        "달력과 명단으로 바로 보여줘요."
    )
    category = st.selectbox("구분", db.CATEGORIES, key="cal_category")
    c1, c2 = st.columns(2)
    with c1:
        target_year = st.number_input(
            "연도", min_value=2020, max_value=2100, value=date.today().year, step=1, key="cal_year"
        )
    with c2:
        target_month = st.number_input(
            "월", min_value=1, max_value=12, value=date.today().month, step=1, key="cal_month"
        )
    st.divider()
    _kakao_calendar_view(category, int(target_year), int(target_month))


def kakao_import_view(user):
    st.subheader("카톡 근태 가져오기")
    st.caption(
        "카톡방(본사/직영/소사장)에서 '대화 내보내기'로 받은 .txt 파일을 통째로 올리면, 그중 아래에서 고른 "
        "달(월)의 근태만 걸러서 정리해드려요. 확인하고 저장하면 '근태표 양식 엑셀 다운로드'에도 바로 반영돼요. "
        "직원들은 지금처럼 카톡에만 올리면 되고, 따로 앱에 입력할 필요 없어요."
    )
    st.caption("💡 이번 달에 어느 날짜를 이미 가져왔는지는 '반영 현황 캘린더' 탭에서 확인할 수 있어요.")

    category = st.selectbox("구분", db.CATEGORIES, key="kakao_category")
    c1, c2 = st.columns(2)
    with c1:
        target_year = st.number_input(
            "가져올 연도", min_value=2020, max_value=2100, value=date.today().year, step=1, key="kakao_year"
        )
    with c2:
        target_month = st.number_input(
            "가져올 월", min_value=1, max_value=12, value=date.today().month, step=1, key="kakao_month"
        )
    st.caption(f"→ {int(target_year)}년 {int(target_month)}월 근태만 걸러서 가져와요. 나머지 달 내용은 자동으로 무시됩니다.")

    uploaded = st.file_uploader(
        "카톡 대화 파일 업로드 (.txt, 카카오톡 '대화 내보내기')", type=["txt"], key="kakao_file"
    )
    chat_text = st.text_area(
        "또는 카톡 대화 내용 직접 붙여넣기", height=180,
        placeholder=f"예시:\n{_KAKAO_EXAMPLES[category]}", key="kakao_text",
    )

    if st.button("미리보기", key="kakao_preview_btn", use_container_width=True):
        source_text = _decode_uploaded_text(uploaded) if uploaded is not None else chat_text
        if not source_text or not source_text.strip():
            st.warning("파일을 올리거나 대화 내용을 붙여넣어주세요.")
            st.session_state.pop("kakao_parsed", None)
        else:
            fallback_date = date(int(target_year), int(target_month), 1)
            if category == "본사":
                all_parsed, unrecognized = parse_headquarters_chat(source_text, fallback_date)
            elif category == "직영":
                all_parsed, unrecognized = parse_jikyeong_chat(source_text, fallback_date)
            else:
                all_parsed, unrecognized = parse_sosajang_chat(source_text, fallback_date)

            parsed = [
                r for r in all_parsed
                if r["work_date"].year == int(target_year) and r["work_date"].month == int(target_month)
            ]
            st.session_state["kakao_parsed"] = parsed
            st.session_state["kakao_parsed_category"] = category
            # 못 읽은 줄도 선택한 연/월에 해당하는 것만 남김
            # (카톡 파일에는 여러 달이 섞여 있어서, 안 그러면 다른 달 공지까지 전부 쌓임)
            st.session_state["kakao_unrecognized"] = [
                ln for ln, d in unrecognized
                if d and d.year == int(target_year) and d.month == int(target_month)
            ]
            st.info(f"전체 인식 {len(all_parsed)}건 중 {int(target_year)}년 {int(target_month)}월 대상 {len(parsed)}건만 가져왔어요.")
            if not parsed:
                st.warning("인식된 근태 내용이 없어요. 형식이 다르면 캡처해서 알려주시면 맞춰드릴게요.")

    parsed = st.session_state.get("kakao_parsed")
    parsed_category = st.session_state.get("kakao_parsed_category")
    unrecognized = st.session_state.get("kakao_unrecognized") or []

    if parsed_category == category and unrecognized:
        with st.expander(f"🔍 확인이 필요할 수도 있는 줄 ({len(unrecognized)}건) — 근태일 수도 있는데 못 읽었어요"):
            st.caption("아래 줄들은 형식이 달라서 자동으로 인식하지 못했어요. 근태 내용이 맞다면 캡처해서 알려주시면 파서를 맞춰드릴게요.")
            for line in unrecognized:
                st.text(line)

    # 저장 결과는 화면을 다시 그린 뒤에 보여줘야 해서 세션에 잠깐 담아뒀다가 여기서 표시
    last = st.session_state.pop("kk_last_result", None)
    if last:
        saved_n, unmatched_n = last
        if saved_n:
            st.success(f"✅ {saved_n}건 저장되었습니다.")
        if unmatched_n:
            st.warning(
                f"직원을 못 찾아서 저장하지 못한 항목이 {unmatched_n}건 있어요. "
                "아래에 그 항목들만 남겨뒀으니, '직원'칸에서 사람을 고른 뒤 다시 저장해주세요."
            )

    if parsed and parsed_category == category:
        st.divider()
        st.markdown(f"**미리보기 — {len(parsed)}건 인식됨. 확인하고 틀린 부분은 고친 뒤 저장하세요.**")

        cat_employees = [e for e in db.list_users(include_inactive=False) if e["category"] == category]
        emp_names = [e["name"] for e in cat_employees]

        def _auto_match(row):
            if category == "소사장" and row.get("store"):
                return next((e for e in cat_employees if row["store"] in e["department"]), None)
            return next((e for e in cat_employees if e["name"] == row.get("name")), None)

        matches = [_auto_match(r) for r in parsed]
        unmatched_idx = [i for i, m in enumerate(matches) if m is None]

        # 저장했더니 매칭 안 된 게 남았으면, 그 건들만 바로 보이도록 자동으로 켜줌
        if st.session_state.pop("kk_focus_unmatched", False):
            st.session_state["kk_only_unmatched"] = True

        only_unmatched = False
        if unmatched_idx:
            st.warning(
                f"⚠️ 직원을 자동으로 못 찾은 항목이 {len(unmatched_idx)}건 있어요. "
                "이름 오타이거나, 계정이 아직 등록되지 않은 사람일 수 있어요."
            )
            only_unmatched = st.checkbox(
                f"매칭 안 된 {len(unmatched_idx)}건만 보기",
                key="kk_only_unmatched",
                help="체크하면 손봐야 하는 것만 보여줘요. 이 상태로 저장하면 보이는 항목만 저장됩니다.",
            )

        visible = unmatched_idx if only_unmatched else list(range(len(parsed)))
        st.caption(f"{len(visible)}건 표시 중 (전체 {len(parsed)}건)")

        with st.form("kakao_save_form"):
            row_keys = []
            for idx in visible:
                row = parsed[idx]
                with st.container(border=True):
                    st.caption(f"원문: {row['raw']}")
                    c1, c2, c3, c4, c5 = st.columns([1.3, 1.6, 1.3, 1.6, 0.8])
                    with c1:
                        st.date_input(
                            "날짜", value=row["work_date"], key=f"kk_date_{idx}", label_visibility="collapsed"
                        )
                    with c2:
                        match = matches[idx]
                        options = ["(직접 선택)"] + emp_names
                        default_idx = options.index(match["name"]) if match and match["name"] in options else 0
                        st.selectbox(
                            "직원", options, index=default_idx, key=f"kk_emp_{idx}", label_visibility="collapsed"
                        )
                    with c3:
                        code_default = row["code"] if row["code"] in db.ATTENDANCE_CODES else "기타"
                        st.selectbox(
                            "근태코드", db.ATTENDANCE_CODES, index=db.ATTENDANCE_CODES.index(code_default),
                            key=f"kk_code_{idx}", label_visibility="collapsed",
                        )
                    with c4:
                        st.text_input(
                            "메모", value=row.get("memo", ""), key=f"kk_memo_{idx}", label_visibility="collapsed"
                        )
                    with c5:
                        st.checkbox("반영", value=True, key=f"kk_include_{idx}")
                    row_keys.append(idx)

            submitted = st.form_submit_button("선택한 내용 저장", use_container_width=True)

        if submitted:
            saved = 0
            unmatched = 0
            for idx in row_keys:
                if not st.session_state.get(f"kk_include_{idx}", True):
                    continue
                chosen_name = st.session_state.get(f"kk_emp_{idx}")
                if chosen_name == "(직접 선택)" or not chosen_name:
                    unmatched += 1
                    continue
                emp = next((e for e in cat_employees if e["name"] == chosen_name), None)
                if not emp:
                    unmatched += 1
                    continue
                wd = st.session_state.get(f"kk_date_{idx}")
                code = st.session_state.get(f"kk_code_{idx}")
                memo = st.session_state.get(f"kk_memo_{idx}", "")
                db.upsert_attendance(emp["id"], wd.isoformat(), code, memo)
                saved += 1
            st.session_state["kk_last_result"] = (saved, unmatched)
            if unmatched:
                # 매칭 안 된 게 남았으면 그 건들만 보이게 화면을 다시 그림
                st.session_state["kk_focus_unmatched"] = True
            else:
                del st.session_state["kakao_parsed"]
                st.session_state.pop("kk_only_unmatched", None)
            st.rerun()


# ---------------------------------------------------------------------------
# 원본 근태 파일 형식(본사근태 / 직영근태 / 소사장근태) 그대로 엑셀 다운로드
# ---------------------------------------------------------------------------

# 본사/직영: 매장(부서) + 직원명 행 x 날짜 열 그리드. 예전 본사근태/직영근태 파일과 동일한 구조.
def _build_roster_matrix_sheet(wb, sheet_name, employees, att_map, dates):
    ws = wb.create_sheet(title=sheet_name)
    weekday_kr = ["월", "화", "수", "목", "금", "토", "일"]
    n = len(dates)

    ws.cell(row=1, column=1, value=f"{sheet_name} 일정표 (빈 칸 = 정상출근)")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6 + n)

    ws.cell(row=2, column=1, value="부서/매장")
    ws.cell(row=2, column=2, value="직원명")
    for i, d in enumerate(dates):
        ws.cell(row=2, column=3 + i, value=weekday_kr[d.weekday()])
    ws.cell(row=2, column=3 + n, value="휴무")
    ws.cell(row=2, column=4 + n, value="연차")
    ws.cell(row=2, column=5 + n, value="반차")
    ws.cell(row=2, column=6 + n, value="합계")

    for i, d in enumerate(dates):
        ws.cell(row=3, column=3 + i, value=d.strftime("%-m/%-d"))

    r = 4
    for e in sorted(employees, key=lambda x: (x["department"], x["name"])):
        ws.cell(row=r, column=1, value=e["department"])
        ws.cell(row=r, column=2, value=e["name"])
        hol = annual = half = 0
        for i, d in enumerate(dates):
            code = att_map.get((e["id"], d.isoformat()))
            ws.cell(row=r, column=3 + i, value="" if code in (None, "정상출근") else code)
            if code == "휴무":
                hol += 1
            elif code == "연차":
                annual += 1
            elif code == "반차":
                half += 1
        ws.cell(row=r, column=3 + n, value=hol)
        ws.cell(row=r, column=4 + n, value=annual)
        ws.cell(row=r, column=5 + n, value=half)
        ws.cell(row=r, column=6 + n, value=hol + annual + half)
        r += 1

    ws.cell(row=r, column=1, value="재직인원")
    for i in range(n):
        ws.cell(row=r, column=3 + i, value=len(employees))


# 소사장: 개인별 근태코드를 예전 '출첵' 파일처럼 매장 단위 요약(o/휴/개인/조기퇴근/지각/미입력)으로 변환
_STORE_STATUS_MAP = {
    "정상출근": "o",
    "특근": "o",
    "당직": "o",
    "교육": "o",
    "휴무": "휴",
    "연차": "개인",
    "반차": "개인",
    "무급": "개인",
    "개인용무": "개인",
    "경조": "개인",
    "예비군": "개인",
    "지각": "지각",
    "조퇴": "조기퇴근",
    "기타": "기타",
}


def _store_status_for_date(codes):
    """같은 매장 소속 직원들의 그날 근태코드 목록 -> 매장 단위 상태 문자열로 요약"""
    if not codes:
        return "미입력"
    statuses = [_STORE_STATUS_MAP.get(c, c) for c in codes]
    if "o" in statuses:
        return "o"
    if all(s == "휴" for s in statuses):
        return "휴"
    uniq = []
    for s in statuses:
        if s not in uniq:
            uniq.append(s)
    return "/".join(uniq)


def _build_store_summary_sheet(wb, employees, att_map, dates):
    ws = wb.create_sheet(title="소사장근태")
    weekday_kr = ["월", "화", "수", "목", "금", "토", "일"]
    n = len(dates)

    seen = set()
    stores = []
    for e in employees:
        if e["department"] not in seen:
            stores.append(e["department"])
            seen.add(e["department"])
    ordered_stores = [s for s in db.SOSAJANG_STORES if s in seen] + [s for s in stores if s not in db.SOSAJANG_STORES]

    ws.cell(
        row=1, column=1,
        value="소사장 매장별 근태 요약 (o=정상, 휴=휴무, 개인=연차·반차 등 개인사유, 조기퇴근, 지각, 미입력=입력 없음)",
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + n)
    ws.cell(row=2, column=1, value="매장명")
    for i, d in enumerate(dates):
        ws.cell(row=2, column=2 + i, value=weekday_kr[d.weekday()])
    for i, d in enumerate(dates):
        ws.cell(row=3, column=2 + i, value=d.strftime("%-m/%-d"))

    emp_by_store = {}
    for e in employees:
        emp_by_store.setdefault(e["department"], []).append(e)

    r = 4
    for store in ordered_stores:
        ws.cell(row=r, column=1, value=store)
        for i, d in enumerate(dates):
            ds = d.isoformat()
            codes = [att_map.get((e["id"], ds)) for e in emp_by_store.get(store, [])]
            codes = [c for c in codes if c]
            ws.cell(row=r, column=2 + i, value=_store_status_for_date(codes))
        r += 1


def build_category_format_excel(start_date, end_date):
    """본사근태 / 직영근태 / 소사장근태 - 예전에 쓰던 파일과 같은 서식으로 엑셀 생성 (구분별 시트 분리)"""
    dates = []
    d = start_date
    while d <= end_date:
        dates.append(d)
        d += timedelta(days=1)
    if not dates:
        return None

    all_employees = db.list_users(include_inactive=False)
    wb = Workbook()
    wb.remove(wb.active)
    made_sheet = False

    for category, sheet_name in [("본사", "본사근태"), ("직영", "직영근태")]:
        emps = [e for e in all_employees if e["category"] == category]
        if not emps:
            continue
        records = db.get_attendance_records_for_matrix(category, start_date.isoformat(), end_date.isoformat())
        att_map = {(r["user_id"], r["work_date"]): r["code"] for r in records}
        _build_roster_matrix_sheet(wb, sheet_name, emps, att_map, dates)
        made_sheet = True

    sosajang_emps = [e for e in all_employees if e["category"] == "소사장"]
    if sosajang_emps:
        records = db.get_attendance_records_for_matrix("소사장", start_date.isoformat(), end_date.isoformat())
        att_map = {(r["user_id"], r["work_date"]): r["code"] for r in records}
        _build_store_summary_sheet(wb, sosajang_emps, att_map, dates)
        made_sheet = True

    if not made_sheet:
        return None
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# 관리자용: 전체 근태현황
# ---------------------------------------------------------------------------
def admin_status_view(user):
    st.subheader("근태 미입력 매장/부서 확인")
    check_date = st.date_input("확인할 날짜", value=date.today(), key="missing_check_date")
    summary = db.get_org_unit_summary_for_date(check_date.isoformat())
    missing = [s for s in summary if s["employee_count"] > 0 and s["entered_count"] == 0]
    if missing:
        st.warning(f"{check_date.isoformat()} 기준, 근태 입력이 하나도 없는 부서/매장이 {len(missing)}곳 있어요.")
        miss_df = pd.DataFrame(missing)[["category", "department", "employee_count"]]
        miss_df.columns = ["구분", "부서/매장", "재직 인원"]
        st.dataframe(miss_df, use_container_width=True, hide_index=True)
    else:
        st.success(f"{check_date.isoformat()} 기준, 모든 부서/매장에 근태가 입력되었어요.")

    st.divider()
    st.subheader("전체 근태 현황")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        start = st.date_input(
            "조회 시작일", value=date.today().replace(day=1), key="admin_start"
        )
    with c2:
        end = st.date_input("조회 종료일", value=date.today(), key="admin_end")
    with c3:
        category = st.selectbox("구분", ["전체"] + db.CATEGORIES, key="admin_category")
    with c4:
        if category == "전체":
            org_options = ["전체"] + db.DEPARTMENTS + db.JIKYEONG_STORES + db.SOSAJANG_STORES
        else:
            org_options = ["전체"] + db.get_org_units(category)
        org_unit = st.selectbox("부서/매장", org_options, key=f"admin_org_{category}")

    st.markdown("**근태표 양식으로 다운로드 (본사·직영·소사장 원본 서식)**")
    st.caption(
        "예전에 쓰시던 본사근태·직영근태 파일과 같은 매장/부서·직원명 × 날짜 표, "
        "소사장근태(출첵) 파일과 같은 매장 단위 요약표로 시트를 나눠서 받아요. "
        "위 조회 시작일/종료일 기준으로 만들어집니다."
    )
    matrix_bytes = build_category_format_excel(start, end)
    if matrix_bytes:
        st.download_button(
            "근태표 양식 엑셀 다운로드",
            data=matrix_bytes,
            file_name=f"근태표_{start.isoformat()}_{end.isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="matrix_download",
        )
    else:
        st.info("등록된 직원이 없어 근태표를 만들 수 없습니다.")

    st.divider()

    records = db.get_all_attendance(start.isoformat(), end.isoformat(), category, org_unit)
    if records:
        df = pd.DataFrame(records)[
            ["work_date", "category", "department", "name", "code", "memo", "updated_at"]
        ]
        df.columns = ["날짜", "구분", "부서/매장", "이름", "근태코드", "메모", "최종수정"]
        st.dataframe(df, use_container_width=True, hide_index=True)

        # 엑셀 다운로드
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="근태현황")
        st.download_button(
            "엑셀로 다운로드",
            data=buf.getvalue(),
            file_name=f"근태현황_{start.isoformat()}_{end.isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.info("해당 조건에 입력된 근태 내역이 없습니다.")


# ---------------------------------------------------------------------------
# 관리자용: 직원 계정 관리
# ---------------------------------------------------------------------------
def admin_accounts_view(user):
    with st.expander("🔀 탭 순서 바꾸기"):
        tab_order_settings_view()

    st.divider()
    st.subheader("직원 계정 추가")

    c1, c2 = st.columns(2)
    with c1:
        add_category = st.selectbox("구분", db.CATEGORIES, key="add_user_category")
    with c2:
        add_org_options = db.get_org_units(add_category)
        add_org_unit = st.selectbox(
            "부서/매장", add_org_options, key=f"add_user_org_{add_category}"
        )

    with st.form("add_user_form"):
        c1, c2 = st.columns(2)
        with c1:
            new_username = st.text_input("아이디 (사번 등, 영문/숫자 권장)")
            new_name = st.text_input("이름")
        with c2:
            new_password = st.text_input("초기 비밀번호", type="password")
        new_role = st.radio("권한", ["employee", "admin"], horizontal=True,
                             format_func=lambda x: "일반 직원" if x == "employee" else "관리자")
        submitted = st.form_submit_button("계정 생성", use_container_width=True)
    if submitted:
        if not new_username or not new_password or not new_name:
            st.error("아이디, 이름, 초기 비밀번호는 필수입니다.")
        else:
            ok, msg = db.create_user(
                new_username.strip(), new_password, new_name.strip(),
                add_category, add_org_unit, new_role,
            )
            if ok:
                st.success(f"{msg} ({add_category} · {add_org_unit})")
                st.rerun()
            else:
                st.error(msg)

    st.divider()
    with st.expander("엑셀로 직원 일괄 등록 (여러 명 한번에)"):
        st.caption(
            "70명을 한 명씩 등록하기 번거로우시면, 아래 양식에 맞춰 엑셀을 채워서 올리세요. "
            "'구분'은 본사/직영/소사장 중 하나, '부서/매장'은 그 구분에 실제로 있는 이름과 정확히 같아야 해요."
        )
        st.download_button(
            "빈 양식 다운로드 (예시 3줄 포함)",
            data=_build_bulk_template(),
            file_name="직원_일괄등록_양식.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="bulk_template_dl",
        )

        if "bulk_upload_key_n" not in st.session_state:
            st.session_state["bulk_upload_key_n"] = 0

        bulk_result = st.session_state.pop("bulk_register_result", None)
        if bulk_result:
            if bulk_result["success"]:
                st.success(f"✅ {bulk_result['success']}명 등록 완료 — 직원 목록에 바로 반영됐어요.")
            if bulk_result["failed"]:
                st.error(f"실패 {len(bulk_result['failed'])}건")
                st.dataframe(
                    pd.DataFrame(bulk_result["failed"], columns=["아이디", "실패 사유"]),
                    use_container_width=True, hide_index=True,
                )

        uploaded = st.file_uploader(
            "작성한 엑셀 파일 업로드 (.xlsx)", type=["xlsx"],
            key=f"bulk_upload_file_{st.session_state['bulk_upload_key_n']}",
        )
        if uploaded is not None:
            try:
                in_df = pd.read_excel(uploaded, dtype=str).fillna("")
            except Exception as e:
                st.error(f"파일을 읽을 수 없습니다: {e}")
                in_df = None

            if in_df is not None:
                required_cols = {"아이디", "이름", "구분", "부서/매장"}
                if not required_cols.issubset(set(in_df.columns)):
                    st.error(f"필수 열이 빠졌습니다. 필요한 열: {', '.join(sorted(required_cols))}")
                else:
                    st.dataframe(in_df, use_container_width=True, hide_index=True)
                    if st.button("일괄 등록 실행", key="bulk_register_btn", use_container_width=True):
                        success, failed = [], []
                        for _, row in in_df.iterrows():
                            uname = str(row.get("아이디", "")).strip()
                            name = str(row.get("이름", "")).strip()
                            category = str(row.get("구분", "")).strip()
                            org_unit = str(row.get("부서/매장", "")).strip()
                            pw = str(row.get("초기비밀번호", "")).strip() or "changeme123"
                            role = str(row.get("권한", "")).strip()
                            role = role if role in ("employee", "admin") else "employee"

                            if not uname or not name:
                                failed.append((uname or "(빈값)", "아이디 또는 이름이 비어있음"))
                                continue
                            if category not in db.CATEGORIES:
                                failed.append((uname, f"구분 '{category}'이 본사/직영/소사장 중 하나가 아님"))
                                continue
                            if org_unit not in db.get_org_units(category):
                                failed.append((uname, f"'{category}'에 '{org_unit}' 부서/매장이 없음"))
                                continue

                            ok, msg = db.create_user(uname, pw, name, category, org_unit, role)
                            (success if ok else failed).append(uname if ok else (uname, msg))

                        st.session_state["bulk_register_result"] = {
                            "success": len(success), "failed": failed,
                        }
                        st.session_state["bulk_upload_key_n"] += 1
                        st.rerun()

    st.divider()
    st.subheader("직원 목록")
    users = db.list_users()
    udf = pd.DataFrame(users)
    if not udf.empty:
        udf_display = udf[["username", "name", "category", "department", "role"]].copy()
        udf_display.columns = ["아이디", "이름", "구분", "부서/매장", "권한"]
        st.dataframe(udf_display, use_container_width=True, hide_index=True)

        with st.expander("복사하기 편한 텍스트로 보기"):
            st.caption("이 표는 그림처럼 그려져서 Ctrl+C가 잘 안 먹혀요. 아래 상자 안 글자는 평범하게 드래그해서 복사하시면 돼요.")
            lines = [
                f"{u['username']}\t{u['name']}\t{u['category']}\t{u['department']}\t{u['role']}"
                for u in users
            ]
            st.text_area(
                "아이디 / 이름 / 구분 / 부서·매장 / 권한",
                value="아이디\t이름\t구분\t부서/매장\t권한\n" + "\n".join(lines),
                height=200, key="user_list_copy_text",
            )

        with st.expander("직원 정보 수정 / 계정 삭제"):
            edit_target = st.selectbox(
                "대상 직원",
                users,
                format_func=lambda u: f"{u['name']} ({u['username']}, {u['category']} · {u['department']})",
                key="edit_target_select",
            )

            e_c1, e_c2 = st.columns(2)
            with e_c1:
                edit_name = st.text_input(
                    "이름", value=edit_target["name"], key=f"edit_name_{edit_target['id']}"
                )
                edit_category = st.selectbox(
                    "구분", db.CATEGORIES,
                    index=db.CATEGORIES.index(edit_target["category"])
                    if edit_target["category"] in db.CATEGORIES else 0,
                    key=f"edit_category_{edit_target['id']}",
                )
            with e_c2:
                edit_org_options = db.get_org_units(edit_category)
                edit_department = st.selectbox(
                    "부서/매장", edit_org_options,
                    index=edit_org_options.index(edit_target["department"])
                    if edit_target["department"] in edit_org_options else 0,
                    key=f"edit_dept_{edit_target['id']}_{edit_category}",
                )
                edit_role = st.radio(
                    "권한", ["employee", "admin"], horizontal=True,
                    index=0 if edit_target["role"] == "employee" else 1,
                    format_func=lambda x: "일반 직원" if x == "employee" else "관리자",
                    key=f"edit_role_{edit_target['id']}",
                )
            edit_new_pw = st.text_input(
                "새 비밀번호 (바꿀 때만 입력, 비워두면 기존 비밀번호 유지)",
                type="password", key=f"edit_pw_{edit_target['id']}",
            )

            if st.button("수정 저장", key=f"edit_save_{edit_target['id']}", use_container_width=True):
                db.update_user(
                    edit_target["id"], edit_name.strip(), edit_category, edit_department,
                    edit_role, edit_new_pw or None,
                )
                st.success(f"{edit_name} 계정 정보가 수정되었습니다.")
                st.rerun()

            st.divider()
            confirm_del = st.checkbox(
                "정말 삭제할게요 (근태 입력 내역도 함께 삭제되며, 되돌릴 수 없어요)",
                key=f"confirm_del_{edit_target['id']}",
            )
            if st.button(
                "계정 완전 삭제", key=f"edit_del_{edit_target['id']}",
                type="secondary", use_container_width=True, disabled=not confirm_del,
            ):
                if edit_target["username"] == "admin":
                    st.error("기본 admin 계정은 삭제할 수 없습니다.")
                else:
                    db.delete_user(edit_target["id"])
                    st.success(f"{edit_target['name']} 계정이 완전히 삭제되었습니다.")
                    st.rerun()

        with st.expander("계정 비활성화(퇴사 처리)"):
            target = st.selectbox(
                "비활성화할 직원",
                users,
                format_func=lambda u: f"{u['name']} ({u['username']}, {u['category']} · {u['department']})",
                key="deactivate_target_select",
            )
            if st.button("선택한 계정 비활성화", type="secondary", key="deactivate_btn"):
                if target["username"] == "admin":
                    st.error("기본 admin 계정은 비활성화할 수 없습니다.")
                else:
                    db.deactivate_user(target["id"])
                    st.success(f"{target['name']} 계정이 비활성화되었습니다.")
                    st.rerun()

        with st.expander("비활성화된 계정 되살리기"):
            inactive_users = [u for u in db.list_users(include_inactive=True) if u["active"] == 0]
            if not inactive_users:
                st.caption("비활성화된 계정이 없습니다.")
            else:
                revive_target = st.selectbox(
                    "되살릴 직원",
                    inactive_users,
                    format_func=lambda u: f"{u['name']} ({u['username']}, {u['category']} · {u['department']})",
                    key="revive_target_select",
                )
                if st.button("선택한 계정 되살리기", key="revive_btn", use_container_width=True):
                    db.activate_user(revive_target["id"])
                    st.success(f"{revive_target['name']} 계정이 다시 활성화되었습니다.")
                    st.rerun()


# ---------------------------------------------------------------------------
# 메인
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# 카톡 AI 검토 — 카톡 대화를 AI가 읽고, 오타·누락·애매한 건을 골라내 알려줌
# ---------------------------------------------------------------------------
def _roster_for(category):
    """AI에게 알려줄 명단 (본사·직영은 직원 이름, 소사장은 매장명)"""
    if category == "소사장":
        return list(db.SOSAJANG_STORES)
    return [
        f"{e['name']} ({e['department']})"
        for e in db.list_users(include_inactive=False)
        if e["category"] == category
    ]


_ISSUE_ICON = {
    "오타": "✏️", "누락": "🕳️", "중복": "♻️", "확인필요": "❓", "규정위반": "🚨",
}


def ai_review_view(user):
    st.subheader("카톡 AI 검토")
    st.caption(
        "카톡 대화를 그대로 올리면 AI가 읽고 ①어떤 근태가 올라왔는지 정리하고 "
        "②오타·안 올라온 사람·중복·애매해서 물어봐야 하는 건을 골라서 알려줘요. "
        "내용을 확인하고 고친 다음 저장하면 근태 현황과 캘린더에 바로 반영돼요."
    )

    if not ai_review.is_configured():
        st.warning(
            "AI 검토를 쓰려면 Anthropic API 키가 필요해요. (앱 코드나 GitHub에는 절대 안 들어갑니다)\n\n"
            "1. console.anthropic.com 에서 API 키를 발급받으세요 (사용한 만큼 과금돼요)\n"
            "2. Streamlit Cloud에서 이 앱 → 우측 상단 ⋮ → **Settings → Secrets** 에 아래 한 줄을 붙여넣고 저장\n"
            "```\nANTHROPIC_API_KEY = \"sk-ant-여기에-발급받은-키\"\n```\n"
            "3. 앱이 자동으로 재시작되면 이 화면에서 바로 쓸 수 있어요."
        )
        st.divider()

    c1, c2, c3 = st.columns([2, 1, 1])
    with c1:
        category = st.selectbox("구분", db.CATEGORIES, key="air_category")
    with c2:
        year = st.number_input("연도", 2020, 2100, date.today().year, 1, key="air_year")
    with c3:
        month = st.number_input("월", 1, 12, date.today().month, 1, key="air_month")

    up = st.file_uploader("카톡 대화 내보내기 파일 (.txt)", type=["txt"], key="air_file")
    text = st.text_area(
        "또는 대화 내용을 여기에 붙여넣기",
        height=160,
        key="air_text",
        placeholder="카톡 대화를 길게 눌러 복사한 뒤 붙여넣어도 돼요.",
    )
    if up is not None:
        text = _decode_uploaded_text(up)
        st.caption(f"📎 {up.name} — {len(text):,}자를 읽었어요.")

    model = db.get_setting("ai_model", ai_review.DEFAULT_MODEL)

    if st.button("🔍 AI 검토 시작", type="primary", disabled=not (text or "").strip()):
        roster = _roster_for(category)
        # 카톡 파일에는 보통 여러 달이 통째로 들어있어요.
        # 고른 달 구간만 잘라서 보내면 요금이 훨씬 적게 나오고 판단도 정확해져요.
        before = len(text)
        text, found = ai_review.filter_month(text, int(year), int(month))
        if found and before > 0:
            st.caption(
                f"📄 {before:,}자 중 {int(year)}년 {int(month)}월 구간 {len(text):,}자만 검토해요"
                f" (약 {100 - int(len(text) / before * 100)}% 절약)"
            )
        if not text.strip():
            st.warning(
                f"올리신 파일에 {int(year)}년 {int(month)}월 대화가 없어요. 연도·월을 확인해주세요."
            )
            st.stop()
        try:
            box = st.empty()

            def _progress(i, n):
                box.info(f"검토 중… ({i}/{n} 조각)" if n > 1 else "검토 중…")

            with st.spinner("AI가 카톡 내용을 읽고 있어요…"):
                result = ai_review.review_kakao(
                    category, text, roster, int(year), int(month),
                    model=model, progress=_progress,
                )
            box.empty()
            rid = db.save_kakao_review(
                category, f"{int(year):04d}-{int(month):02d}", ai_review.to_json(result)
            )
            st.session_state["air_result"] = result
            st.session_state["air_review_id"] = rid
        except Exception as e:
            st.error(f"검토 중 문제가 생겼어요.\n\n{e}")

    result = st.session_state.get("air_result")
    if not result:
        st.info("아직 검토한 내용이 없어요. 카톡 대화를 올리고 위 버튼을 눌러주세요.")
        _recent_reviews_view()
        return

    st.divider()
    # 지난 검토를 불러온 경우 위쪽 선택칸과 다를 수 있으므로, 결과 자체에 적힌 구분·월을 따른다
    res_category = result.get("category") or category
    res_year = int(result.get("year") or year)
    res_month = int(result.get("month") or month)
    st.markdown(f"#### 📄 {res_category} · {res_year}년 {res_month}월 검토 결과")
    if result.get("summary"):
        st.markdown(f"**📌 검토 요약**\n\n{result['summary']}")

    issues = result.get("issues") or []
    st.markdown(f"### ⚠️ 확인이 필요한 것 ({len(issues)}건)")
    if not issues:
        st.success("이상한 점은 발견되지 않았어요.")
    else:
        for kind in ["확인필요", "오타", "누락", "중복", "규정위반"]:
            group = [i for i in issues if i.get("kind") == kind]
            if not group:
                continue
            with st.expander(f"{_ISSUE_ICON.get(kind,'•')} {kind} ({len(group)}건)", expanded=(kind == "확인필요")):
                for it in group:
                    head = " ".join(x for x in [it.get("date", ""), it.get("target", "")] if x)
                    st.markdown(f"- **{head}** {it.get('detail','')}")
                    if it.get("suggestion"):
                        st.caption(f"　→ {it['suggestion']}")

    records = result.get("records") or []
    st.markdown(f"### 📋 읽어낸 근태 내용 ({len(records)}건)")
    st.caption(
        "표를 직접 고칠 수 있어요. 잘못된 줄은 맨 왼쪽 칸을 체크 해제하면 저장에서 빠져요. "
        "다 고치셨으면 아래 저장 버튼을 눌러주세요."
    )
    if not records:
        st.info("카톡에서 확인된 근태 기록이 없어요.")
        return

    is_sosajang = (res_category == "소사장")
    rows = []
    for r in records:
        rows.append({
            "저장": (r.get("confidence") != "low"),
            "날짜": r.get("date", ""),
            ("매장" if is_sosajang else "이름"): r.get("target", ""),
            ("출근보고" if is_sosajang else "근태"): r.get("code", ""),
            **({"퇴근보고": r.get("close_code", "")} if is_sosajang else {}),
            "메모": r.get("memo", ""),
            "확신도": r.get("confidence", ""),
            "근거": r.get("source", ""),
        })
    edited = st.data_editor(
        pd.DataFrame(rows),
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        key="air_editor",
        column_config={
            "저장": st.column_config.CheckboxColumn("저장", width="small"),
            "확신도": st.column_config.TextColumn("확신도", disabled=True, width="small"),
            "근거": st.column_config.TextColumn("근거", disabled=True),
        },
    )

    if st.button("💾 확인한 내용 저장하기", type="primary"):
        saved, failed = 0, []
        for _, row in edited.iterrows():
            if not row.get("저장"):
                continue
            dstr = str(row.get("날짜", "")).strip()
            target = str(row.get("매장" if is_sosajang else "이름", "")).strip()
            code = str(row.get("출근보고" if is_sosajang else "근태", "")).strip()
            memo = str(row.get("메모", "") or "").strip()
            if not dstr or not target:
                continue
            if is_sosajang:
                close_code = str(row.get("퇴근보고", "") or "").strip()
                db.upsert_store_attendance(
                    store=target, work_date=dstr, open_code=code,
                    close_code=close_code,
                    perf_code=("O" if code.lower() == "o" else code),
                    memo=memo,
                )
                saved += 1
            else:
                matches = db.find_user_by_name(target, res_category)
                if not matches:
                    failed.append(f"{dstr} {target} (등록된 계정 없음)")
                    continue
                if len(matches) > 1:
                    failed.append(f"{dstr} {target} (동명이인 {len(matches)}명 — 직접 입력해주세요)")
                    continue
                db.upsert_attendance(matches[0]["id"], dstr, code, memo)
                saved += 1

        if st.session_state.get("air_review_id"):
            db.update_kakao_review(st.session_state["air_review_id"], status="applied")
        st.success(f"{saved}건 저장했어요. '반영 현황 캘린더' 탭에서 바로 확인할 수 있어요.")
        if failed:
            st.warning("아래 항목은 저장하지 못했어요:\n\n" + "\n".join(f"- {f}" for f in failed))

    _recent_reviews_view()


def _recent_reviews_view():
    reviews = db.list_kakao_reviews(limit=10)
    if not reviews:
        return
    with st.expander("🕘 지난 검토 기록 다시 보기"):
        for rv in reviews:
            mark = "✅ 반영됨" if rv["status"] == "applied" else "⏳ 미반영"
            cols = st.columns([3, 1])
            with cols[0]:
                st.markdown(f"**{rv['category']} · {rv['target_month']}** {mark}")
                st.caption(rv["created_at"][:16].replace("T", " "))
            with cols[1]:
                if st.button("불러오기", key=f"air_load_{rv['id']}"):
                    full = db.get_kakao_review(rv["id"])
                    st.session_state["air_result"] = ai_review.from_json(full["result_json"])
                    st.session_state["air_review_id"] = rv["id"]
                    st.rerun()


# ---------------------------------------------------------------------------
# 엑셀 내보내기 — 근태 양식 4종 + 출퇴근 현황 파일 기재
# ---------------------------------------------------------------------------
def _month_bounds(year, month):
    start = date(year, month, 1)
    end = date(year, 12, 31) if month == 12 else date(year, month + 1, 1) - timedelta(days=1)
    return start, end


def _person_records(category, start, end):
    """{(이름, 'YYYY-MM-DD'): 코드}, {(이름, 날짜): 메모}"""
    rows = db.get_all_attendance(start.isoformat(), end.isoformat(), category=category)
    recs = {(r["name"], r["work_date"]): r["code"] for r in rows}
    memos = {(r["name"], r["work_date"]): (r["memo"] or "") for r in rows if r.get("memo")}
    return recs, memos


def _store_records(start, end):
    """{(매장, 'YYYY-MM-DD'): {'open','close','perf','memo'}}"""
    out = {}
    for r in db.get_store_attendance(start.isoformat(), end.isoformat()):
        out[(r["store"], r["work_date"])] = {
            "open": r["open_code"] or "",
            "close": r["close_code"] or "",
            "perf": r["perf_code"] or "",
            "memo": r["memo"] or "",
        }
    return out


def _report_caption(rep):
    bits = [f"{rep.get('written', 0)}칸 기재"]
    if rep.get("skipped_rows"):
        bits.append(f"⚠️ 양식 줄 수 부족으로 {rep['skipped_rows']}명 누락")
    if rep.get("gaps_dropped"):
        bits.append("인원이 많아 매장 사이 빈 줄은 생략")
    um = rep.get("unmatched") or []
    if um:
        bits.append(f"⚠️ 양식에서 못 찾은 대상 {len(um)}건")
    st.caption(" · ".join(bits))
    if um:
        with st.expander(f"못 찾은 {len(um)}건 보기"):
            st.write(um[:100])
    if rep.get("unknown_codes"):
        with st.expander(f"규정표에 없어 메모로만 넣은 코드 {len(rep['unknown_codes'])}건"):
            st.write(rep["unknown_codes"][:100])


def excel_export_view(user):
    st.subheader("엑셀 내보내기")
    st.caption(
        "DB에 저장된 근태를 실제 쓰시는 엑셀 양식에 그대로 채워서 내려받는 곳이에요. "
        "수식은 건드리지 않고 값만 채워 넣습니다."
    )

    c1, c2 = st.columns(2)
    with c1:
        year = st.number_input("연도", 2020, 2100, date.today().year, 1, key="exp_year")
    with c2:
        month = st.number_input("월", 1, 12, date.today().month, 1, key="exp_month")
    year, month = int(year), int(month)
    start, end = _month_bounds(year, month)
    stamp = f"{year}년_{month:02d}월"

    st.divider()
    st.markdown("### 1️⃣ 근태 양식 4종")
    st.caption("빈 양식에 이번 달 근태를 채워서 바로 내려받아요. 명단은 '직원 계정관리'에 등록된 사람 기준이에요.")

    # --- 본사 근태 양식 ---
    with st.container(border=True):
        st.markdown("**본사 근태 양식**")
        if st.button("만들기", key="exp_bonsa"):
            roster = [
                {"name": e["name"], "department": e["department"]}
                for e in db.list_users(include_inactive=False) if e["category"] == "본사"
            ]
            roster.sort(key=lambda e: (db.DEPARTMENTS.index(e["department"])
                                       if e["department"] in db.DEPARTMENTS else 99, e["name"]))
            recs, _ = _person_records("본사", start, end)
            try:
                data, rep = excel_export.build_bonsa_form(year, month, roster, recs)
                st.session_state["exp_bonsa_out"] = (data, rep)
            except Exception as e:
                st.error(str(e))
        if st.session_state.get("exp_bonsa_out"):
            data, rep = st.session_state["exp_bonsa_out"]
            _report_caption(rep)
            st.download_button("⬇️ 본사 근태 양식 내려받기", data,
                               file_name=f"{stamp}_본사_근태.xlsx", key="exp_bonsa_dl",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # --- 직영점 근태 양식 ---
    with st.container(border=True):
        st.markdown("**직영점 근태 양식**")
        if st.button("만들기", key="exp_jik"):
            roster = [
                {"name": e["name"], "department": e["department"]}
                for e in db.list_users(include_inactive=False) if e["category"] == "직영"
            ]
            recs, _ = _person_records("직영", start, end)
            try:
                data, rep = excel_export.build_jikyeong_form(
                    year, month, roster, recs, store_order=db.JIKYEONG_STORES
                )
                st.session_state["exp_jik_out"] = (data, rep)
            except Exception as e:
                st.error(str(e))
        if st.session_state.get("exp_jik_out"):
            data, rep = st.session_state["exp_jik_out"]
            _report_caption(rep)
            st.download_button("⬇️ 직영점 근태 양식 내려받기", data,
                               file_name=f"{stamp}_직영점_근태.xlsx", key="exp_jik_dl",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # --- 소사장 근태 + 실적보고 ---
    with st.container(border=True):
        st.markdown("**소사장 근태 양식 · 소사장 실적보고 양식**")
        st.caption("소사장은 매장 단위라 하루에 출근보고/퇴근보고 두 칸이 들어가요.")
        if st.button("만들기", key="exp_sosa"):
            sr = _store_records(start, end)
            try:
                d1, r1 = excel_export.build_sosajang_form(year, month, sr)
                d2, r2 = excel_export.build_sosajang_perf_form(year, month, sr)
                st.session_state["exp_sosa_out"] = (d1, r1, d2, r2)
            except Exception as e:
                st.error(str(e))
        if st.session_state.get("exp_sosa_out"):
            d1, r1, d2, r2 = st.session_state["exp_sosa_out"]
            _report_caption(r1)
            st.download_button("⬇️ 소사장 근태 양식 내려받기", d1,
                               file_name=f"{stamp}_소사장_근태.xlsx", key="exp_sosa_dl1",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            st.download_button("⬇️ 소사장 실적보고 양식 내려받기", d2,
                               file_name=f"{stamp}_소사장_실적보고.xlsx", key="exp_sosa_dl2",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.divider()
    st.markdown("### 2️⃣ 출퇴근 현황 파일 채우기")
    st.caption(
        "쓰시던 월간 출퇴근 현황 파일을 그대로 올리면, '▶ 데이터' 탭의 해당 열만 채워서 돌려드려요. "
        "나머지 시트·수식·지문 기록은 손대지 않아요."
    )

    # --- 본사·직영 → G~U ---
    with st.container(border=True):
        st.markdown("**출퇴근 현황 (본사·직영) → ▶ 데이터 탭 G~U열**")
        f1 = st.file_uploader("파일 올리기 (.xlsx)", type=["xlsx"], key="exp_st_bj")
        clear1 = st.checkbox("기존에 적혀 있던 G~U 값은 지우고 새로 채우기", True, key="exp_st_bj_clear")
        if f1 is not None and st.button("채우기", key="exp_st_bj_run"):
            recs_b, memo_b = _person_records("본사", start, end)
            recs_j, memo_j = _person_records("직영", start, end)
            recs = {**recs_b, **recs_j}
            memos = {**memo_b, **memo_j}
            try:
                with st.spinner("파일을 채우는 중이에요…"):
                    data, rep = excel_export.fill_status_bonsa_jikyeong(
                        f1.getvalue(), year, month, recs, memos, clear_existing=clear1
                    )
                st.session_state["exp_st_bj_out"] = (data, rep, f1.name)
            except Exception as e:
                st.error(str(e))
        if st.session_state.get("exp_st_bj_out"):
            data, rep, fname = st.session_state["exp_st_bj_out"]
            st.caption(f"데이터 탭 {rep.get('rows',0):,}행 중 {rep.get('written',0)}칸 기재")
            _report_caption(rep)
            st.download_button("⬇️ 채워진 파일 내려받기", data, file_name=fname, key="exp_st_bj_dl",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # --- 소사장 → K~AH ---
    with st.container(border=True):
        st.markdown("**출퇴근 현황 (소사장) → ▶ 데이터 탭 K~AH열**")
        st.info(
            "이 파일은 원래 .xls라서 그대로는 수식이 깨져요. "
            "엑셀에서 열고 **[다른 이름으로 저장] → 파일 형식을 'Excel 통합 문서(*.xlsx)'** 로 저장한 뒤 올려주세요. "
            "채워진 파일을 받으시면 다시 .xls로 저장하시면 됩니다."
        )
        f2 = st.file_uploader("파일 올리기 (.xlsx)", type=["xlsx"], key="exp_st_s")
        clear2 = st.checkbox("기존에 적혀 있던 K~AH 값은 지우고 새로 채우기", True, key="exp_st_s_clear")
        unrep = st.checkbox("출첵 기준으로 출근/퇴근 미보고(N·O열)도 자동 체크", False, key="exp_st_s_unrep")
        if f2 is not None and st.button("채우기", key="exp_st_s_run"):
            sr = _store_records(start, end)
            try:
                with st.spinner("파일을 채우는 중이에요…"):
                    data, rep = excel_export.fill_status_sosajang(
                        f2.getvalue(), year, month, sr,
                        mark_unreported=unrep, clear_existing=clear2,
                    )
                st.session_state["exp_st_s_out"] = (data, rep, f2.name)
            except Exception as e:
                st.error(str(e))
        if st.session_state.get("exp_st_s_out"):
            data, rep, fname = st.session_state["exp_st_s_out"]
            st.caption(f"데이터 탭 {rep.get('rows',0):,}행 중 {rep.get('written',0)}칸 기재")
            _report_caption(rep)
            st.download_button("⬇️ 채워진 파일 내려받기", data, file_name=fname, key="exp_st_s_dl",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def main():
    if "user" not in st.session_state:
        login_view()
        return

    user = st.session_state["user"]
    logout_button()

    st.title("🗂️ (주)대교통신 HR 플랫폼")

    view_funcs = {
        "ai": lambda: ai_review_view(user),
        "export": lambda: excel_export_view(user),
        "bulk": lambda: bulk_entry_view(user),
        "emp": lambda: employee_view(user),
        "status": lambda: admin_status_view(user),
        "kakao": lambda: kakao_import_view(user),
        "calendar": lambda: kakao_calendar_view(user),
        "admin": lambda: admin_accounts_view(user),
    }

    if user["role"] == "admin":
        order = _get_tab_order()
    else:
        # 일반 직원은 근태일괄입력·근태개별입력 2개만, 관리자가 정한 순서를 그대로 따름
        order = [k for k in _get_tab_order() if k in ("bulk", "emp")]
        if not order:
            order = ["bulk", "emp"]

    tabs = st.tabs([TAB_LABELS[k] for k in order])
    for tab, key in zip(tabs, order):
        with tab:
            view_funcs[key]()


if __name__ == "__main__":
    main()
