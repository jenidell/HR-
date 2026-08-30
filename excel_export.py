"""
excel_export.py
DB에 저장된 근태를 실제 업무용 엑셀 양식에 채워주는 모듈.

두 가지 일을 합니다.

(1) 근태 양식 4종 만들기 — templates/ 안의 빈 양식에 DB 내용을 채워서 내려줌
      · 본사 근태 양식        (본사근태 시트, D6:AH34)
      · 직영점 근태 양식      (입력하는곳 시트, C6:AG21)
      · 소사장 근태 양식      (출첵 시트, 하루 2칸 = 출근보고/퇴근보고)
      · 소사장 실적보고 양식  (실적_월 시트, 하루 1칸)

(2) 출퇴근 현황 파일 채우기 — 사용자가 올린 월간 파일의 '▶ 데이터' 탭을 채워서 돌려줌
      · 출퇴근 현황(본사·직영) → G~U열
      · 출퇴근 현황(소사장)    → K~AH열

수식은 그대로 보존합니다 (openpyxl을 data_only=False로 열어 편집).
"""

import io
import re
import zipfile
import xml.etree.ElementTree as ET
from datetime import date, datetime

import openpyxl
from openpyxl.utils import get_column_letter

# 빈 양식 4종은 templates_data.py 안에 글자로 들어있습니다 (폴더로 올릴 필요 없음)
import templates_data


# ---------------------------------------------------------------------------
# 공통 도우미
# ---------------------------------------------------------------------------

def _tpl(key):
    return openpyxl.load_workbook(io.BytesIO(templates_data.get_template_bytes(key)))


def _save(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def norm_store(s):
    """매장명 표기 흔들림 흡수: '대교대리점 군포역점' / '군포역점' / '군포역' → '군포역'"""
    s = str(s or "").strip()
    s = re.sub(r"^대교대리점\s*", "", s)
    s = s.replace(" ", "")
    if s.endswith("점"):
        s = s[:-1]
    return s


def _match_store(target, candidates_norm):
    """정규화 후 정확히 맞는 게 없으면 앞부분이 겹치는 매장으로 한 번 더 시도"""
    t = norm_store(target)
    if t in candidates_norm:
        return t
    for c in candidates_norm:
        if c.startswith(t) or t.startswith(c):
            return c
    return None


def _to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip().replace(".", "-").replace("/", "-")
        m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s)
        if m:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return None


def _days_in_month(year, month):
    import calendar
    return calendar.monthrange(year, month)[1]


def _display_code(code, memo=""):
    """
    근태 항목 목록에 없는 표현은 '기타'로 저장되는데, 그대로 쓰면 원래 뭐였는지 알 수 없습니다.
    그래서 엑셀에 적을 때는 '기타(휴가)'처럼 원문을 괄호에 남깁니다.
    """
    code = str(code or "").strip()
    memo = str(memo or "").strip()
    if code == "기타" and memo:
        return f"기타({memo})"
    return code


# 근태 양식에서 '정상 출근'은 빈칸으로 둡니다 (COUNTIFS가 특이사항만 세기 때문)
_BLANK_CODES = {"정상출근", "출", "o", "O", ""}


# ---------------------------------------------------------------------------
# (1-1) 본사 근태 양식
# ---------------------------------------------------------------------------

BONSA_FIRST_ROW, BONSA_LAST_ROW = 6, 34



# ---------------------------------------------------------------------------
# (1-2) 직영점 근태 양식
# ---------------------------------------------------------------------------

JIK_FIRST_ROW, JIK_LAST_ROW = 6, 21
JIK_SUMMARY_FIRST_ROW = 3  # 요약 시트 L3부터 명단


def _ensure_jikyeong_row_formulas(ws, r):
    """직영 양식에서 비어있던 행을 쓰게 될 때 집계 수식을 채워 넣음"""
    if not ws.cell(row=r, column=1).value:
        ws.cell(row=r, column=1).value = "=IFERROR(VLOOKUP(B:B,요약!L:M,2,0),0)"
    pairs = {
        35: f"=COUNTIFS(C{r}:AG{r},AI$5)",      # AI 휴무
        36: f"=COUNTIFS(C{r}:AG{r},AJ$5)",      # AJ 연차
        37: f"=COUNTIFS(C{r}:AG{r},AK$5)",      # AK 휴가
        38: f"=$AK$3-AI{r}",                    # AL 특근
        39: f'=COUNTIFS(C{r}:AG{r},"지각")',     # AM 지각
        40: f"=SUM(AI{r}:AK{r})",               # AN 합계
    }
    for c, f in pairs.items():
        if not ws.cell(row=r, column=c).value:
            ws.cell(row=r, column=c).value = f



# ---------------------------------------------------------------------------
# (1-3) 소사장 근태 양식 (출첵) — 하루 2칸
# ---------------------------------------------------------------------------

SOSA_FIRST_ROW, SOSA_LAST_ROW = 4, 15


def sosajang_day_col(day):
    """출첵 시트에서 해당 날짜의 '출근보고' 칸 열 번호 (퇴근보고는 +1)"""
    if day <= 15:
        return 2 + 2 * (day - 1)      # B,D,F ... AD
    return 33 + 2 * (day - 16)        # AG,AI ... BK



# ---------------------------------------------------------------------------
# (1-4) 소사장 실적보고 양식 (실적_월) — 하루 1칸
# ---------------------------------------------------------------------------

PERF_FIRST_ROW, PERF_LAST_ROW = 3, 19



# ---------------------------------------------------------------------------
# (2-1) 출퇴근 현황(본사·직영) — '▶ 데이터' 탭 G~U열
# ---------------------------------------------------------------------------

# 근태코드 → (열번호, 채울 값).  G(소명건)·P·R·S는 지문 대조용이라 건드리지 않습니다.
STATUS_BJ_MAP = {
    "휴무":   (8,  1),     # H
    "연차":   (9,  1),     # I
    "반차":   (10, 0.5),   # J
    "예비군": (11, 1),     # K
    "지각":   (12, 1),     # L
    "경조":   (13, 1),     # M
    "교육":   (14, 1),     # N
    "특근":   (15, 1),     # O
    "무급":   (17, 0.5),   # Q
    "당직":   (20, 1),     # T
}
STATUS_BJ_MEMO_COL = 21   # U
STATUS_BJ_SHEET = "▶ 데이터"
STATUS_BJ_HEADER_ROW = 6
STATUS_BJ_NAME_COL = 30   # AD
STATUS_BJ_DATE_COL = 31   # AE


STATUS_BJ_MEMO_COL = 21   # U


def _build_index(file_bytes, sheet_name, key_col_a, key_col_b, first_row):
    """
    ▶데이터 탭에서 (키A, 키B) → 행번호 색인을 만든다.
    read_only 모드 + 빈 줄 200개 넘어가면 중단 → 큰 파일도 1초 안에 끝남.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        names = ", ".join(wb.sheetnames)
        wb.close()
        raise ValueError(f"'{sheet_name}' 탭을 찾을 수 없습니다. 이 파일의 탭 목록: {names}")
    ws = wb[sheet_name]
    rows, blanks = [], 0
    lo, hi = min(key_col_a, key_col_b), max(key_col_a, key_col_b)
    for i, row in enumerate(
        ws.iter_rows(min_row=first_row, min_col=lo, max_col=hi, values_only=True),
        start=first_row,
    ):
        a, b = row[key_col_a - lo], row[key_col_b - lo]
        if a in (None, "") and b in (None, ""):
            blanks += 1
            if blanks > 200:
                break
            continue
        blanks = 0
        rows.append((i, a, b))
    wb.close()
    return rows


# ---------------------------------------------------------------------------
# (2-1) 출퇴근 현황(본사·직영) — '▶ 데이터' 탭 G~U열
# ---------------------------------------------------------------------------

def fill_status_bonsa_jikyeong(file_bytes, year, month, records, memos=None,
                               clear_existing=True):
    """
    올린 '출퇴근 현황(본사·직영)' 파일의 ▶데이터 탭 G~U열을 DB 기준으로 채웁니다.
    records : {(이름, 'YYYY-MM-DD'): 근태코드}
    memos   : {(이름, 'YYYY-MM-DD'): 메모}
    """
    memos = memos or {}
    raw = _build_index(file_bytes, STATUS_BJ_SHEET,
                       STATUS_BJ_NAME_COL, STATUS_BJ_DATE_COL,
                       STATUS_BJ_HEADER_ROW + 1)
    index = {}
    for i, name, dv in raw:
        d = _to_date(dv)
        if name and d:
            index[(str(name).strip(), d)] = i
    if not index:
        raise ValueError(
            "▶데이터 탭에서 이름(AD열)·일자(AE열)를 읽지 못했습니다. "
            "엑셀에서 파일을 한 번 열었다가 저장한 뒤 다시 올려주세요."
        )

    report = {"written": 0, "unmatched": [], "rows": len(index)}
    updates = {}

    if clear_existing:
        for r in set(index.values()):
            for c, _ in STATUS_BJ_MAP.values():
                updates[(r, c)] = None
            updates[(r, STATUS_BJ_MEMO_COL)] = None

    for (name, dstr), code in records.items():
        d = _to_date(dstr)
        if not d or d.year != year or d.month != month:
            continue
        r = index.get((str(name).strip(), d))
        if r is None:
            report["unmatched"].append(f"{dstr} {name}")
            continue
        hit = STATUS_BJ_MAP.get(code)
        if hit:
            updates[(r, hit[0])] = hit[1]
            report["written"] += 1
        memo = memos.get((name, dstr)) or ""
        if not hit and code not in _BLANK_CODES:
            # 표에 없는 코드는 임의로 버리지 말고 메모로 남겨 담당자가 판단하게 함
            memo = (code + (" / " + memo if memo else "")).strip()
        if memo:
            updates[(r, STATUS_BJ_MEMO_COL)] = memo

    out, missing_rows = inject(file_bytes, STATUS_BJ_SHEET, updates)
    report["rows_not_found"] = missing_rows
    return out, report


# ---------------------------------------------------------------------------
# (2-2) 출퇴근 현황(소사장) — '▶ 데이터' 탭 K~AH열
# ---------------------------------------------------------------------------

STATUS_S_SHEET = "▶ 데이터"
STATUS_S_HEADER_ROW = 5
STATUS_S_DATE_COL = 35    # AI 발생일자
STATUS_S_STORE_COL = 36   # AJ 매장명
STATUS_S_MEMO_COL = 11    # K 메모

# 출첵 코드 → (K열 메모, 체크(=1)할 열 목록)
STATUS_S_MAP = {
    "휴":       ("휴무",     [20]),          # T 휴무
    "휴무":     ("휴무",     [20]),
    "미":       ("휴무",     [20, 17]),      # T 휴무 + Q 매장미오픈
    "조기마감": ("조기마감", [29]),          # AC
    "조기퇴근": ("조기퇴근", [30]),          # AD
    "개인":     ("개인용무", [31]),          # AE
    "개인용무": ("개인용무", [31]),
    "휴가":     ("휴가",     [21]),          # U
    "본사":     ("본사회의", []),            # 정보성 메모만 (결근 아님)
    "본사회의": ("본사회의", []),
}
STATUS_S_ALL_CHECK_COLS = [17, 20, 21, 29, 30, 31]


def fill_status_sosajang(file_bytes, year, month, store_records,
                         mark_unreported=False, clear_existing=True):
    """
    올린 '출퇴근 현황(소사장)' 파일의 ▶데이터 탭 K~AH열을 DB 기준으로 채웁니다.
    store_records : {(매장명, 'YYYY-MM-DD'): {'open':..., 'close':..., 'memo':...}}
    mark_unreported=True 면 출첵 기준으로 출근/퇴근 미보고(N/O열)도 함께 체크합니다.
    """
    raw = _build_index(file_bytes, STATUS_S_SHEET,
                       STATUS_S_DATE_COL, STATUS_S_STORE_COL,
                       STATUS_S_HEADER_ROW + 1)
    index, store_names = {}, set()
    for i, dv, store in raw:
        d = _to_date(dv)
        if store and d:
            n = norm_store(store)
            index[(n, d)] = i
            store_names.add(n)
    if not index:
        raise ValueError(
            "▶데이터 탭에서 발생일자(AI열)·매장명(AJ열)을 읽지 못했습니다. "
            "엑셀에서 파일을 한 번 열었다가 저장한 뒤 다시 올려주세요."
        )

    report = {"written": 0, "unmatched": [], "unknown_codes": [], "rows": len(index)}
    updates = {}

    if clear_existing:
        for r in set(index.values()):
            updates[(r, STATUS_S_MEMO_COL)] = None
            for c in STATUS_S_ALL_CHECK_COLS:
                updates[(r, c)] = None
            if mark_unreported:
                updates[(r, 14)] = None  # N 출근미보고
                updates[(r, 15)] = None  # O 퇴근미보고

    cand = sorted(store_names)
    for (store, dstr), vals in store_records.items():
        key = _match_store(store, cand)
        d = _to_date(dstr)
        if key is None or not d or d.year != year or d.month != month:
            report["unmatched"].append(f"{dstr} {store}")
            continue
        r = index.get((key, d))
        if r is None:
            report["unmatched"].append(f"{dstr} {store}")
            continue

        open_code = (vals.get("open") or "").strip()
        close_code = (vals.get("close") or "").strip()
        extra_memo = (vals.get("memo") or "").strip()

        memo_parts, cols = [], []
        for part in [p.strip() for p in re.split(r"[/+,]", open_code) if p.strip()]:
            if part in ("o", "O"):
                continue
            hit = STATUS_S_MAP.get(part)
            if hit:
                memo_parts.append(hit[0])
                cols.extend(hit[1])
            else:
                # 규정표에 없는 문구는 원문 그대로 메모로 남김 (임의 판단 금지)
                memo_parts.append(part)
                report["unknown_codes"].append(f"{dstr} {store}: {part}")
        if extra_memo:
            memo_parts.append(extra_memo)

        if memo_parts:
            updates[(r, STATUS_S_MEMO_COL)] = "/".join(dict.fromkeys(memo_parts))
        for c in dict.fromkeys(cols):
            updates[(r, c)] = 1

        if mark_unreported:
            is_off = any(m in ("휴무", "휴가") for m in memo_parts)
            if not open_code:
                updates[(r, 14)] = 1
            elif not close_code and not is_off:
                updates[(r, 15)] = 1

        if memo_parts or cols:
            report["written"] += 1

    out, missing_rows = inject(file_bytes, STATUS_S_SHEET, updates)
    report["rows_not_found"] = missing_rows
    return out, report


# ---------------------------------------------------------------------------
# 엑셀 파일에 값만 콕 집어 써넣기 (메모리 절약용)
#
# xlsx는 사실 zip 파일입니다. 값을 써야 하는 시트 하나의 XML만 훑으면서 고치고
# 나머지는 바이트 그대로 복사합니다. openpyxl로 통째로 열었다 저장하면
# 큰 파일에서 1GB 가까이 써서 Streamlit Cloud에서 앱이 죽기 때문입니다.
# ---------------------------------------------------------------------------
NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
ET.register_namespace("", NS)


def col_letter(idx):
    """1 → A, 27 → AA"""
    s = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        s = chr(65 + rem) + s
    return s


def col_index(letters):
    """A → 1, AA → 27"""
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch.upper()) - 64)
    return n


def _sheet_xml_path(zf, sheet_name):
    """시트 이름 → zip 안의 xl/worksheets/sheetN.xml 경로"""
    wb = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rid_to_target = {}
    for rel in rels:
        target = rel.get("Target")
        if target.startswith("/"):
            target = target[1:]
        elif not target.startswith("xl/"):
            target = "xl/" + target
        rid_to_target[rel.get("Id")] = target.replace("xl/worksheets/../", "xl/")
    for sheet in wb.find(f"{{{NS}}}sheets"):
        if sheet.get("name") == sheet_name:
            rid = sheet.get(f"{{{NS_R}}}id")
            return rid_to_target.get(rid)
    return None


def _make_cell(ref, value, style=None):
    """<c> 엘리먼트 하나 생성. 문자열은 inlineStr로 넣어서 sharedStrings를 안 건드림."""
    c = ET.Element(f"{{{NS}}}c", {"r": ref})
    if style is not None:
        c.set("s", style)
    if value is None or value == "":
        return c
    if isinstance(value, str) and value.startswith("="):
        # '='로 시작하면 글자가 아니라 수식으로 넣어야 합니다.
        # (<is>로 넣으면 엑셀이 수식 문자열을 그대로 보여줍니다)
        f = ET.SubElement(c, f"{{{NS}}}f")
        f.text = value[1:]
    elif isinstance(value, bool):
        c.set("t", "b")
        v = ET.SubElement(c, f"{{{NS}}}v")
        v.text = "1" if value else "0"
    elif isinstance(value, (int, float)):
        v = ET.SubElement(c, f"{{{NS}}}v")
        v.text = repr(value) if isinstance(value, float) else str(value)
    else:
        c.set("t", "inlineStr")
        is_el = ET.SubElement(c, f"{{{NS}}}is")
        t = ET.SubElement(is_el, f"{{{NS}}}t")
        t.text = str(value)
        t.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    return c


_NSDECL_RE = re.compile(rb'xmlns(?::[A-Za-z0-9_.-]+)?="[^"]*"')


def _rewrite_row(row_xml, updates_for_row, extra_ns=b""):
    """
    행 하나(<row>...</row>)의 XML을 받아 지정된 열들을 갱신한 XML 바이트로 반환.
    extra_ns : 워크시트 최상단에 선언된 네임스페이스들(x14ac 등).
               행 안에 x14ac:dyDescent 같은 속성이 있으면 이게 없으면 파싱이 깨집니다.
    """
    wrapper = b'<r xmlns="' + NS.encode() + b'" ' + extra_ns + b">" + row_xml + b"</r>"
    root = ET.fromstring(wrapper)
    row = root[0]

    existing = {}
    for c in list(row):
        ref = c.get("r") or ""
        m = re.match(r"([A-Z]+)(\d+)", ref)
        if m:
            existing[col_index(m.group(1))] = c

    row_no = row.get("r")
    for cidx, value in updates_for_row.items():
        ref = f"{col_letter(cidx)}{row_no}"
        old = existing.get(cidx)
        style = old.get("s") if old is not None else None
        new_c = _make_cell(ref, value, style)
        if old is not None:
            row.remove(old)
        existing[cidx] = new_c

    # 열 순서대로 다시 배치 (엑셀은 셀이 열 순서대로 정렬돼 있어야 함)
    for c in list(row):
        row.remove(c)
    for cidx in sorted(existing):
        c = existing[cidx]
        # 값도 수식도 없는 빈 셀은 굳이 남기지 않음
        if len(c) == 0 and c.get("t") is None and c.get("s") is None:
            continue
        row.append(c)

    out = ET.tostring(row, encoding="utf-8")
    # ElementTree가 다시 붙이는 네임스페이스 선언 제거 (워크시트 최상단에 이미 선언돼 있음)
    out = _NSDECL_RE.sub(b"", out)
    return out.replace(b"<row  ", b"<row ")


_ROW_RE = re.compile(rb"<row\b[^>]*/>|<row\b[^>]*>.*?</row>", re.S)
_ROW_NUM_RE = re.compile(rb'<row[^>]*\br="(\d+)"')


def inject(xlsx_bytes, sheet_name, updates):
    """
    updates : {(행번호, 열번호): 값}   값이 None이면 해당 셀을 비움
    반환    : 수정된 xlsx 바이트
    """
    by_row = {}
    for (r, c), v in updates.items():
        by_row.setdefault(int(r), {})[int(c)] = v
    if not by_row:
        return xlsx_bytes

    src = zipfile.ZipFile(io.BytesIO(xlsx_bytes))
    target = _sheet_xml_path(src, sheet_name)
    if not target:
        src.close()
        raise ValueError(f"'{sheet_name}' 시트를 찾을 수 없습니다.")

    data = src.read(target)

    # 워크시트 최상단의 네임스페이스 선언을 그대로 가져다 씁니다.
    # (행 속성에 x14ac:dyDescent 같은 게 있으면 이게 없으면 XML 파싱이 실패)
    extra_ns, root_tag = b"", re.search(rb"<worksheet\b[^>]*>", data)
    if root_tag:
        decls = re.findall(rb'xmlns:[A-Za-z0-9_.-]+="[^"]*"', root_tag.group(0))
        extra_ns = b" ".join(decls)
        for pref, uri in re.findall(rb'xmlns:([A-Za-z0-9_.-]+)="([^"]*)"', root_tag.group(0)):
            ET.register_namespace(pref.decode(), uri.decode())

    start = data.find(b"<sheetData")
    if start < 0:
        src.close()
        raise ValueError("시트 XML 구조를 해석하지 못했습니다.")
    body_start = data.find(b">", start) + 1
    end = data.find(b"</sheetData>")
    if end < 0:  # <sheetData/> 형태 (내용 없음)
        src.close()
        raise ValueError("시트에 데이터 영역이 없습니다.")

    out = [data[:body_start]]
    pos = body_start
    touched = set()
    for m in _ROW_RE.finditer(data, body_start, end):
        num_m = _ROW_NUM_RE.match(m.group(0))
        if not num_m:
            continue
        rno = int(num_m.group(1))
        if rno not in by_row:
            continue
        out.append(data[pos:m.start()])
        out.append(_rewrite_row(m.group(0), by_row[rno], extra_ns))
        pos = m.end()
        touched.add(rno)
    out.append(data[pos:end])
    out.append(data[end:])
    new_sheet = b"".join(out)
    del data, out

    buf = io.BytesIO()
    dst = zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED)
    for item in src.infolist():
        if item.filename == target:
            dst.writestr(item, new_sheet)
        else:
            dst.writestr(item, src.read(item.filename))
    dst.close()
    src.close()
    buf.seek(0)
    return buf.getvalue(), sorted(set(by_row) - touched)


# ===========================================================================
# 올려주신 "내 양식 파일"에 값만 써넣기 (서식·수식·메모·이미지·인쇄설정 100% 보존)
#
# openpyxl로 열었다 저장하면 셀 메모·도형·인쇄설정 같은 게 사라집니다.
# 그래서 원본 파일은 손대지 않고, 값이 들어갈 칸만 XML 수준에서 갈아끼웁니다.
# ===========================================================================

def _sheet_or_fail(file_bytes, candidates):
    """양식 파일에서 기대하는 시트를 찾음. 이름이 다르면 어떤 시트가 있는지 알려줌."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    names = wb.sheetnames
    wb.close()
    for c in candidates:
        if c in names:
            return c
    raise ValueError(
        f"'{candidates[0]}' 시트를 찾을 수 없습니다. 이 파일의 시트: {', '.join(names)}"
    )


def _read_keys(file_bytes, sheet, first_row, last_row, col):
    """양식 안의 명단/매장명을 읽어 {이름: 행번호}로 반환 (수식이면 계산된 값을 읽음)"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb[sheet]
    out = {}
    for i, row in enumerate(
        ws.iter_rows(min_row=first_row, max_row=last_row, min_col=col, max_col=col, values_only=True),
        start=first_row,
    ):
        v = row[0]
        if v not in (None, ""):
            out[str(v).strip()] = i
    wb.close()
    return out


def _clear(updates, rows, cols):
    for r in rows:
        for c in cols:
            updates[(r, c)] = None


def fill_form_bonsa(file_bytes, year, month, records, memos=None, clear_existing=True):
    """올린 본사 근태 양식(본사근태 시트 D6:AH34)에 근태를 채웁니다."""
    sheet = _sheet_or_fail(file_bytes, ["본사근태"])
    row_of = _read_keys(file_bytes, sheet, BONSA_FIRST_ROW, BONSA_LAST_ROW, 3)  # C열 직원명
    if not row_of:
        raise ValueError("양식의 C6:C34에서 직원명을 읽지 못했습니다. 명단이 비어있는 파일인지 확인해주세요.")

    updates = {(3, 2): year, (3, 3): month, (3, 36): year}   # B3=연, C3=월, AJ3=연
    updates.update(BONSA_DAY_UPDATES())   # 날짜·요일 줄이 그 달에 맞게 자동으로 바뀌도록
    if clear_existing:
        _clear(updates, range(BONSA_FIRST_ROW, BONSA_LAST_ROW + 1), range(4, 35))

    report = {"written": 0, "unmatched": [], "roster": len(row_of)}
    ndays = _days_in_month(year, month)
    for (name, dstr), code in records.items():
        r = row_of.get(str(name).strip())
        if r is None:
            report["unmatched"].append(f"{dstr} {name}")
            continue
        d = _to_date(dstr)
        if not d or d.year != year or d.month != month or d.day > ndays:
            continue
        if code in _BLANK_CODES:
            continue
        updates[(r, 3 + d.day)] = _display_code(code, (memos or {}).get((str(name).strip(), dstr)))
        report["written"] += 1

    out, _ = inject(file_bytes, sheet, updates)
    return out, report


def fill_form_jikyeong(file_bytes, year, month, records, memos=None, clear_existing=True):
    """올린 직영점 근태 양식(입력하는곳 시트 C6:AG21)에 근태를 채웁니다."""
    sheet = _sheet_or_fail(file_bytes, ["입력하는곳"])
    row_of = _read_keys(file_bytes, sheet, JIK_FIRST_ROW, JIK_LAST_ROW, 2)  # B열 직원명(수식→계산값)
    if not row_of:
        raise ValueError(
            "양식의 B6:B21에서 직원명을 읽지 못했습니다. "
            "엑셀에서 파일을 한 번 열었다가 저장한 뒤 다시 올려주세요."
        )

    updates = {(3, 2): month, (3, 31): year}   # B3=월, AE3=연
    updates.update(JIK_DAY_UPDATES())     # 날짜·요일 줄이 그 달에 맞게 자동으로 바뀌도록
    if clear_existing:
        _clear(updates, sorted(row_of.values()), range(3, 34))

    report = {"written": 0, "unmatched": [], "roster": len(row_of)}
    ndays = _days_in_month(year, month)
    for (name, dstr), code in records.items():
        r = row_of.get(str(name).strip())
        if r is None:
            report["unmatched"].append(f"{dstr} {name}")
            continue
        d = _to_date(dstr)
        if not d or d.year != year or d.month != month or d.day > ndays:
            continue
        if code in _BLANK_CODES and code != "출":
            continue
        updates[(r, 2 + d.day)] = _display_code(code, (memos or {}).get((str(name).strip(), dstr)))
        report["written"] += 1

    out, _ = inject(file_bytes, sheet, updates)
    return out, report


def fill_form_sosajang(file_bytes, year, month, store_records, clear_existing=True):
    """올린 소사장 근태 양식(출첵 시트)에 채웁니다. 하루 2칸(출근보고/퇴근보고)."""
    sheet = _sheet_or_fail(file_bytes, ["출첵"])
    raw = _read_keys(file_bytes, sheet, SOSA_FIRST_ROW, SOSA_LAST_ROW, 1)
    row_of = {norm_store(k): v for k, v in raw.items()}
    if not row_of:
        raise ValueError("양식의 A4:A15에서 매장명을 읽지 못했습니다.")

    updates = {(3, 1): f"{month}월"}
    if clear_existing:
        cols = []
        for day in range(1, 32):
            c = sosajang_day_col(day)
            cols += [c, c + 1]
        _clear(updates, sorted(row_of.values()), cols)

    report = {"written": 0, "unmatched": [], "roster": len(row_of)}
    cand = sorted(row_of)
    ndays = _days_in_month(year, month)
    for (store, dstr), vals in store_records.items():
        key = _match_store(store, cand)
        if key is None:
            report["unmatched"].append(f"{dstr} {store}")
            continue
        d = _to_date(dstr)
        if not d or d.year != year or d.month != month or d.day > ndays:
            continue
        r, c = row_of[key], sosajang_day_col(d.day)
        open_code = (vals.get("open") or "").strip()
        close_code = (vals.get("close") or "").strip()
        memo = (vals.get("memo") or "").strip()
        updates[(r, c)] = memo if memo and not open_code else (open_code or None)
        updates[(r, c + 1)] = close_code or None
        if open_code or close_code or memo:
            report["written"] += 1

    out, _ = inject(file_bytes, sheet, updates)
    return out, report


def fill_form_sosajang_perf(file_bytes, year, month, store_records, clear_existing=True):
    """올린 소사장 실적보고 양식(실적_월 시트 B3:AF19)에 채웁니다."""
    sheet = _sheet_or_fail(file_bytes, ["실적_월"])
    raw = _read_keys(file_bytes, sheet, PERF_FIRST_ROW, PERF_LAST_ROW, 1)
    row_of = {norm_store(k): v for k, v in raw.items()}
    if not row_of:
        raise ValueError("양식의 A3:A19에서 매장명을 읽지 못했습니다.")

    updates = {(1, 1): f"{month}월"}
    if clear_existing:
        _clear(updates, sorted(row_of.values()), range(2, 33))

    report = {"written": 0, "unmatched": [], "roster": len(row_of)}
    cand = sorted(row_of)
    ndays = _days_in_month(year, month)
    for (store, dstr), vals in store_records.items():
        key = _match_store(store, cand)
        if key is None:
            report["unmatched"].append(f"{dstr} {store}")
            continue
        d = _to_date(dstr)
        if not d or d.year != year or d.month != month or d.day > ndays:
            continue
        val = (vals.get("perf") or "").strip()
        memo = (vals.get("memo") or "").strip()
        if not val and not memo:
            continue
        updates[(row_of[key], 1 + d.day)] = memo if memo and not val else val
        report["written"] += 1

    out, _ = inject(file_bytes, sheet, updates)
    return out, report


# ===========================================================================
# 근태 양식(채워진 것)을 거꾸로 읽어오기
#
# "근태 양식에 기재된 내용을 기준으로 출퇴근 현황에 반영" 하기 위한 것.
# 양식을 내려받아 손으로 고치셨다면, 그 고친 내용이 그대로 반영됩니다.
# ===========================================================================

def _parse_display_code(v):
    """'기타(재택근무)' → ('기타', '재택근무') / '휴무' → ('휴무', '')"""
    s = str(v or "").strip()
    if not s:
        return "", ""
    m = re.match(r"^(.+?)\s*\((.+)\)\s*$", s)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return s, ""


def _read_grid(file_bytes, sheet, first_row, last_row, key_col, day_to_col, year, month):
    """행=대상, 열=날짜인 격자를 읽어 {(대상, 'YYYY-MM-DD'): (코드, 메모)} 로 반환"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        names = ", ".join(wb.sheetnames)
        wb.close()
        raise ValueError(f"'{sheet}' 시트를 찾을 수 없습니다. 이 파일의 시트: {names}")
    ws = wb[sheet]
    ndays = _days_in_month(year, month)
    max_col = max(day_to_col(d) for d in range(1, ndays + 1)) + 1
    out = {}
    for i, row in enumerate(
        ws.iter_rows(min_row=first_row, max_row=last_row, min_col=1, max_col=max_col, values_only=True),
        start=first_row,
    ):
        key = row[key_col - 1] if len(row) >= key_col else None
        if key in (None, ""):
            continue
        key = str(key).strip()
        for day in range(1, ndays + 1):
            c = day_to_col(day)
            v = row[c - 1] if len(row) >= c else None
            code, memo = _parse_display_code(v)
            if code:
                out[(key, f"{year:04d}-{month:02d}-{day:02d}")] = (code, memo)
    wb.close()
    return out


def read_form_bonsa(file_bytes, year, month):
    """본사 근태 양식 → {(이름, 날짜): 코드}, {(이름, 날짜): 메모}"""
    sheet = _sheet_or_fail(file_bytes, ["본사근태"])
    grid = _read_grid(file_bytes, sheet, BONSA_FIRST_ROW, BONSA_LAST_ROW, 3,
                      lambda d: 3 + d, year, month)
    return ({k: v[0] for k, v in grid.items()},
            {k: v[1] for k, v in grid.items() if v[1]})


def read_form_jikyeong(file_bytes, year, month):
    """직영점 근태 양식 → {(이름, 날짜): 코드}, {(이름, 날짜): 메모}"""
    sheet = _sheet_or_fail(file_bytes, ["입력하는곳"])
    grid = _read_grid(file_bytes, sheet, JIK_FIRST_ROW, JIK_LAST_ROW, 2,
                      lambda d: 2 + d, year, month)
    return ({k: v[0] for k, v in grid.items()},
            {k: v[1] for k, v in grid.items() if v[1]})


def read_form_sosajang(file_bytes, year, month):
    """소사장 근태 양식(출첵) → {(매장, 날짜): {'open','close','memo'}}"""
    sheet = _sheet_or_fail(file_bytes, ["출첵"])
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb[sheet]
    ndays = _days_in_month(year, month)
    out = {}
    for i, row in enumerate(
        ws.iter_rows(min_row=SOSA_FIRST_ROW, max_row=SOSA_LAST_ROW,
                     min_col=1, max_col=sosajang_day_col(31) + 1, values_only=True),
        start=SOSA_FIRST_ROW,
    ):
        store = row[0]
        if store in (None, ""):
            continue
        store = str(store).strip()
        for day in range(1, ndays + 1):
            c = sosajang_day_col(day)
            o = str(row[c - 1] or "").strip() if len(row) >= c else ""
            cl = str(row[c] or "").strip() if len(row) > c else ""
            if not o and not cl:
                continue
            key = f"{year:04d}-{month:02d}-{day:02d}"
            # 규정표에 있는 코드가 아니면 메모(예외 문구)로 간주
            known = o in ("o", "O", "휴", "휴무", "미", "개인", "휴가", "본사",
                          "조기마감", "조기퇴근") or "/" in o
            out[(store, key)] = {
                "open": o if known else "",
                "close": cl,
                "memo": "" if known else o,
                "perf": ("O" if o in ("o", "O") else o),
            }
    wb.close()
    return out


# ===========================================================================
# 날짜 줄을 '그 달에 맞게' 자동으로 바뀌도록 만들기
#
# 양식의 날짜 줄(1~31)이 고정 숫자라, 30일까지인 달에도 31일 칸이 남아 있었습니다.
# 아래 수식으로 바꾸면 연/월만 바꿔도 없는 날짜는 자동으로 비고, 요일도 따라 바뀝니다.
#   DATE(연, 월+1, 0) = 그 달의 마지막 날  →  DAY(...) = 그 달의 일수
# ===========================================================================

def _day_row_updates(day_first_col, day_row, weekday_row, year_ref, month_ref, ndays_max=31):
    """날짜 줄과 요일 줄을 '연/월에 따라 자동으로 바뀌는 수식'으로 교체할 내용을 만듭니다."""
    up = {}
    for day in range(1, ndays_max + 1):
        c = day_first_col + (day - 1)
        cl = get_column_letter(c)
        up[(day_row, c)] = (
            f"=IF({day}>DAY(DATE({year_ref},{month_ref}+1,0)),\"\",{day})"
        )
        if weekday_row:
            up[(weekday_row, c)] = (
                f'=IF({cl}{day_row}="","",'
                f'TEXT(WEEKDAY(DATE({year_ref},{month_ref},{cl}{day_row}),1),"AAA"))'
            )
    return up


# 본사 근태 양식: 날짜 D5:AH5, 요일 D4:AH4, 연=$AJ$3, 월=$C$3
BONSA_DAY_UPDATES = lambda: _day_row_updates(4, 5, 4, "$AJ$3", "$C$3")
# 직영점 근태 양식: 날짜 C5:AG5, 요일 C4:AG4, 연=$AE$3, 월=$B$3
JIK_DAY_UPDATES = lambda: _day_row_updates(3, 5, 4, "$AE$3", "$B$3")


def build_bonsa_form(year, month, roster, records, memos=None):
    """앱에 들어있는 본사 양식에 명단과 근태를 채웁니다 (서식·수식·메모 100% 보존)."""
    data = templates_data.get_template_bytes("본사")
    report = {"written": 0, "skipped_rows": 0, "unmatched": []}

    capacity = BONSA_LAST_ROW - BONSA_FIRST_ROW + 1
    if len(roster) > capacity:
        report["skipped_rows"] = len(roster) - capacity
        roster = roster[:capacity]

    updates = {(3, 2): year, (3, 3): month, (3, 36): year}
    updates.update(BONSA_DAY_UPDATES())
    row_of = {}
    for i, emp in enumerate(roster):
        r = BONSA_FIRST_ROW + i
        updates[(r, 2)] = emp.get("department") or ""
        updates[(r, 3)] = emp.get("name") or ""
        row_of[emp.get("name")] = r
    for r in range(BONSA_FIRST_ROW + len(roster), BONSA_LAST_ROW + 1):
        updates[(r, 2)] = None
        updates[(r, 3)] = None

    ndays = _days_in_month(year, month)
    for (name, dstr), code in records.items():
        r = row_of.get(name)
        if r is None:
            report["unmatched"].append(f"{dstr} {name}")
            continue
        d = _to_date(dstr)
        if not d or d.year != year or d.month != month or d.day > ndays:
            continue
        if code in _BLANK_CODES:
            continue
        updates[(r, 3 + d.day)] = _display_code(code, (memos or {}).get((name, dstr)))
        report["written"] += 1

    out, _ = inject(data, "본사근태", updates)
    return out, report


def build_jikyeong_form(year, month, roster, records, memos=None, store_order=None):
    """앱에 들어있는 직영점 양식에 명단과 근태를 채웁니다 (서식·수식·메모 100% 보존)."""
    data = templates_data.get_template_bytes("직영")
    report = {"written": 0, "skipped_rows": 0, "unmatched": []}

    groups = {}
    for emp in roster:
        groups.setdefault(emp.get("department") or "", []).append(emp)
    order = [s for s in (store_order or []) if s in groups]
    order += [s for s in groups if s not in order]

    # 1) 요약 시트에 명단 기록 (입력하는곳 B열 수식이 여기를 참조)
    sum_up, sum_row, summary_row_of = {}, JIK_SUMMARY_FIRST_ROW, {}
    for store in order:
        for emp in groups[store]:
            sum_up[(sum_row, 12)] = emp.get("name")
            sum_up[(sum_row, 13)] = store
            summary_row_of[emp.get("name")] = sum_row
            sum_row += 1
    for r in range(sum_row, JIK_SUMMARY_FIRST_ROW + 20):
        sum_up[(r, 12)] = None
        sum_up[(r, 13)] = None
    data, _ = inject(data, "요약", sum_up)

    # 2) 입력하는곳 배치 (자리가 모자라면 매장 사이 빈 줄 생략)
    capacity = JIK_LAST_ROW - JIK_FIRST_ROW + 1
    total_people = sum(len(groups[s]) for s in order)
    use_gaps = (total_people + max(0, len(order) - 1)) <= capacity
    report["gaps_dropped"] = not use_gaps

    updates = {(3, 2): month, (3, 31): year}
    updates.update(JIK_DAY_UPDATES())
    r, row_of = JIK_FIRST_ROW, {}
    for gi, store in enumerate(order):
        if gi > 0 and use_gaps:
            r += 1
        for emp in groups[store]:
            if r > JIK_LAST_ROW:
                report["skipped_rows"] += 1
                continue
            updates[(r, 1)] = "=IFERROR(VLOOKUP(B:B,요약!L:M,2,0),0)"
            updates[(r, 2)] = f"=요약!L{summary_row_of[emp['name']]}"
            updates[(r, 35)] = f"=COUNTIFS(C{r}:AG{r},AI$5)"
            updates[(r, 36)] = f"=COUNTIFS(C{r}:AG{r},AJ$5)"
            updates[(r, 37)] = f"=COUNTIFS(C{r}:AG{r},AK$5)"
            updates[(r, 38)] = f"=$AK$3-AI{r}"
            updates[(r, 39)] = f'=COUNTIFS(C{r}:AG{r},"지각")'
            updates[(r, 40)] = f"=SUM(AI{r}:AK{r})"
            row_of[emp["name"]] = r
            r += 1
    for rr in range(r, JIK_LAST_ROW + 1):
        updates[(rr, 2)] = None

    ndays = _days_in_month(year, month)
    for (name, dstr), code in records.items():
        rr = row_of.get(name)
        if rr is None:
            report["unmatched"].append(f"{dstr} {name}")
            continue
        d = _to_date(dstr)
        if not d or d.year != year or d.month != month or d.day > ndays:
            continue
        if code in _BLANK_CODES and code != "출":
            continue
        updates[(rr, 2 + d.day)] = _display_code(code, (memos or {}).get((name, dstr)))
        report["written"] += 1

    out, _ = inject(data, "입력하는곳", updates)
    return out, report


def build_sosajang_form(year, month, store_records):
    """앱에 들어있는 소사장 근태 양식(출첵)에 채웁니다. 매장명은 양식에 이미 있습니다."""
    return fill_form_sosajang(templates_data.get_template_bytes("소사장"),
                              year, month, store_records)


def build_sosajang_perf_form(year, month, store_records):
    """앱에 들어있는 소사장 실적보고 양식(실적_월)에 채웁니다."""
    return fill_form_sosajang_perf(templates_data.get_template_bytes("소사장실적"),
                                   year, month, store_records)

